# -*- coding: utf-8 -*-
"""
Created on Wed Jul 25 12:32:08 2018

Pyene Outputs provides methods for saving simualtion outputs in HDF5 files

@author: Dr Eduardo Alejandro Martínez Ceseña
https://www.researchgate.net/profile/Eduardo_Alejandro_Martinez_Cesena
"""
import numpy as np
from tables import Int16Col, Float32Col, StringCol, IsDescription, open_file
import os
from pyomo.core import ConcreteModel
from .pyene_Models import EnergyandNetwork, Networkmodel, characteristic
from openpyxl import Workbook

class pyeneOConfig:
    ''' Default settings used for this class '''
    def __init__(self):
        # Default time-step and map
        self.settings = {}
        self.settings['Directory1'] = None  # Location of main folder
        self.settings['Directory2'] = None  # Location of main folder
        self.settings['Name1'] = 'pyeneOutputs.h5'  # Name
        self.settings['Name2'] = 'pyeneSensitivity.h5'  # Name
        self.settings['Name3'] = 'pyeneDetailedOutputs.h5'  # Name
        self.settings['Case'] = 2  # 0: None, 1: Single, 2: Sensitivity study

        self.data = {}
        self.data['name'] = None
        self.data['cost'] = [0, 0, 0, 0, 0]
        self.data['curtailment'] = 0
        self.data['spill'] = 0
        self.data['OF'] = 0
        self.data['model'] = None

        self.time = {}
        self.time['All'] = 0
        self.time['glpk'] = 0
        self.time['step'] = None


class pyeneSave():

    def initialise(self, model=None, dir=None, name_file=None):
        self.save_h5 = False
        self.save_excel = False
        for out in model.data.outputs:
            if out.exist("data device") and out.get_characteristic("data device") == ["h5"] and not self.save_h5:
                self.save_h5 = True
                self.fileh5 = pyeneHDF5Settings()
                self.fileh5.initialise(dir, name_file)
            elif out.exist("data device") and out.get_characteristic("data device") == ["excel"] and not self.save_excel:
                self.save_excel = True
                self.fileexcel = pyene2excel()
                self.fileexcel.initialise(dir, name_file)
                
    def save_results(self, model=None, sim_no=None):
        if self.save_h5:
            self.fileh5.save_results(model, sim_no)
        if self.save_excel:
            self.fileexcel.save_results(model, sim_no)
    
    def close_output_files(self):
        self.fileh5.terminate(new_implementation=True)

class pyene2excel():

    def initialise(self, dir=None, name_file=None):
        self.file2save_address = os.path.join(dir, name_file)
    
    def save_results(self, model=None, sim_no=None):
        from .pyene_Models import models
        if isinstance(model, models):
            values, starts, characteristics = model.get_outputs()
        new_file2save = self.file2save_address + 'Simulation_{:05d}.xlsx'.format(sim_no)
        
        wb = Workbook()
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))

        # Extracting name to create subgroups
        information_name = []
        information_problem = []
        information_position_tree = []
        for out in model.data.outputs:
            if out.exist("data device") and out.get_characteristic("data device") == ["excel"]:
                information_name.append(out.get_characteristic("name")[0])
                information_problem.append(out.get_characteristic("problem")[0])
                #TODO: Generalise subscripts
                subscripts = model.get_latest_subscripts()
                for key, value in subscripts.items():
                    if key == "pt":
                        information_position_tree.append(value[1])
                # if out.exist("pt"):
                #     information_position_tree.append(out.get_characteristic("pt"))
                # else:
                #     information_position_tree.append(None)

        used_values = [False for _ in range(len(values))]
        for i_n, i_p, i_pt in zip(information_name, information_problem, information_position_tree):
            if (i_p == "DC OPF"):
                table_columns = {}
                aux = 1
                for cha_group in characteristics:
                    is_group = False
                    for cha in cha_group:
                        if cha[0].decode('utf-8') == "name" and cha[2].decode('utf-8') == i_n:
                            is_group = True
                    if is_group:
                        for cha in cha_group:
                            if cha[0].decode('utf-8') == "pt" or cha[0].decode('utf-8') == "hour":
                                table_columns[cha[0].decode('utf-8')] = aux
                                aux = aux + 1
                        break
                # Adding IDs
                for cha_group in characteristics:
                    is_group = False
                    for cha in cha_group:
                        if cha[0].decode('utf-8') == "name" and cha[2].decode('utf-8') == i_n:
                            is_group = True
                    if is_group:
                        for cha in cha_group:
                            if cha[0].decode('utf-8') == "ID" and cha[2].decode('utf-8') not in table_columns:
                                table_columns[cha[2].decode('utf-8')] = aux
                                aux = aux + 1
                                break
                
                if (i_pt is not None):
                    word = "_" + i_pt[0]
                    for c in i_pt[1:]:
                        word = word + "." + c
                else:
                    word = ""
                
                ws = wb.create_sheet()
                ws.cell(row=1, column=1, value=i_n)
                for key, val in table_columns.items():
                    ws.cell(row=2, column=val, value=key)

                aux = 3
                is_end = False
                while not is_end:
                    is_end = True
                    if (False not in used_values): break
                    reference = {}
                    for num, (val, u_val, cha_group) in enumerate(zip(values, used_values, characteristics)):
                        is_group = True
                        if not u_val:
                            for cha in cha_group:
                                if (cha[0].decode('utf-8') == "name" and cha[2].decode('utf-8') != i_n):
                                    is_group = False
                                if (cha[0].decode('utf-8') == "pt"):
                                    if (len(cha[2:]) != len(i_pt)): is_group = False
                                    else:
                                        for num1, c in enumerate(cha[2:]):
                                            if (c.decode('utf-8') != i_pt[num1]):
                                                is_group = False
                                                break
                        if is_group and not u_val and not reference:
                            is_end = False
                            for cha in cha_group:
                                if (cha[0].decode('utf-8') == "pt" or cha[0].decode('utf-8') == "hour") and len(cha) == 3:
                                    reference[cha[0].decode('utf-8')] = cha[2].decode('utf-8')
                                    ws.cell(row=aux, column=table_columns[cha[0].decode('utf-8')], value=cha[2].decode('utf-8'))
                                elif (cha[0].decode('utf-8') == "pt" or cha[0].decode('utf-8') == "hour") and len(cha) > 3:
                                    word = cha[2].decode('utf-8')
                                    for c in cha[3:]:
                                        word = word + "|" + c.decode('utf-8')
                                    reference[cha[0].decode('utf-8')] = word
                                    ws.cell(row=aux, column=table_columns[cha[0].decode('utf-8')], value=word)
                                elif cha[0].decode('utf-8') == "ID":
                                    ws.cell(row=aux, column=table_columns[cha[2].decode('utf-8')], value=val)
                                    used_values[num] = True
                        elif is_group and not u_val and reference:
                            compare_cha = {}
                            ID = ""
                            for cha in cha_group:
                                if (cha[0].decode('utf-8') == "pt" or cha[0].decode('utf-8') == "hour") and len(cha) == 3:
                                    compare_cha[cha[0].decode('utf-8')] = cha[2].decode('utf-8')
                                elif (cha[0].decode('utf-8') == "pt" or cha[0].decode('utf-8') == "hour") and len(cha) > 3:
                                    word = cha[2].decode('utf-8')
                                    for c in cha[3:]:
                                        word = word + "|" + c.decode('utf-8')
                                    compare_cha[cha[0].decode('utf-8')] = word
                                elif cha[0].decode('utf-8') == "ID":
                                    ID = cha[2].decode('utf-8')
                            if compare_cha == reference:
                                is_end = False
                                ws.cell(row=aux, column=table_columns[ID], value=val)
                                used_values[num] = True
                    aux = aux + 1
            if (i_p == "BT"):
                table_columns = {}
                aux = 1
                for cha_group in characteristics:
                    is_group = False
                    for cha in cha_group:
                        if cha[0].decode('utf-8') == "name" and cha[2].decode('utf-8') == i_n:
                            is_group = True
                    if is_group:
                        for cha in cha_group:
                            if cha[0].decode('utf-8') != "name" and cha[0].decode('utf-8') != "reference":
                                table_columns[cha[0].decode('utf-8')] = aux
                                aux = aux + 1
                # Adding references
                for cha_group in characteristics:
                    is_group = False
                    for cha in cha_group:
                        if cha[0].decode('utf-8') == "name" and cha[2].decode('utf-8') == i_n:
                            is_group = True
                    if is_group:
                        for cha in cha_group:
                            if cha[0].decode('utf-8') == "reference" and cha[2].decode('utf-8') not in table_columns:
                                table_columns[cha[2].decode('utf-8')] = aux
                                aux = aux + 1
                                break
                
                ws = wb.create_sheet(i_n)
                ws.cell(row=1, column=1, value=i_n)
                for key, val in table_columns:
                    ws.cell(row=1, column=val, value=key)
                
                aux = 2
                is_end = False
                while not is_end:
                    is_end = True
                    reference = {}
                    if (False not in used_values): break
                    for num, (val, u_val, cha_group) in enumerate(zip(values, used_values, characteristics)):
                        is_group = False
                        if not u_val:
                            for cha in cha_group:
                                if cha[0].decode('utf-8') == "name" and cha[2].decode('utf-8') == i_n:
                                    is_group = True
                        if is_group and not u_val and not reference:
                            is_end = False
                            for cha in cha_group:
                                if cha[0].decode('utf-8') != "name" and cha[0].decode('utf-8') != "reference" and len(cha) == 3:
                                    reference[cha[0].decode('utf-8')] = cha[2].decode('utf-8')
                                    ws.cell(row=aux, column=table_columns[cha[0].decode('utf-8')], value=cha[2].decode('utf-8'))
                                elif cha[0].decode('utf-8') != "name" and cha[0].decode('utf-8') != "reference" and len(cha) > 3:
                                    word = cha[2].decode('utf-8')
                                    for c in cha[3:]:
                                        word = word + "|" + c.decode('utf-8')
                                    reference[cha[0].decode('utf-8')] = word
                                    ws.cell(row=aux, column=table_columns[cha[0].decode('utf-8')], value=word)
                                elif cha[0].decode('utf-8') == "reference":
                                    ws.cell(row=aux, column=table_columns[cha[2].decode('utf-8')], value=val)
                                    used_values[num] = True
                        elif is_group and not u_val and reference:
                            compare_cha = {}
                            ID = ""
                            for cha in cha_group:
                                if cha[0].decode('utf-8') != "name" and cha[0].decode('utf-8') != "reference" and len(cha) == 3:
                                    compare_cha[cha[0].decode('utf-8')] = cha[2].decode('utf-8')
                                elif cha[0].decode('utf-8') != "name" and cha[0].decode('utf-8') != "reference" and len(cha) > 3:
                                    word = cha[2].decode('utf-8')
                                    for c in cha[3:]:
                                        word = word + "|" + c.decode('utf-8')
                                    compare_cha[cha[0].decode('utf-8')] = word
                                elif cha[0].decode('utf-8') == "reference":
                                    ID = cha[2].decode('utf-8')
                            if compare_cha == reference:
                                is_end = False
                                ws.cell(row=aux, column=table_columns[ID], value=val)
                                used_values[num] = True
                    aux = aux + 1

        wb.save(new_file2save)


class pyeneHDF5Settings():
    def __init__(self, obj=None):
        ''' Initialise network class '''
        # Get default values
        if obj is None:
            obj = pyeneOConfig()

        # Copy attributes
        for pars in obj.__dict__.keys():
            setattr(self, pars, getattr(obj, pars))

        # No outputs if there is no output directory
        if self.settings['Directory1'] is not None:
            self.settings['Name1'] = os.path.join(self.settings['Directory1'],
                                                  self.settings['Name1'])
            self.fileh = open_file(self.settings['Name1'], mode='w')
            self.settings['Name3'] = os.path.join(self.settings['Directory1'],
                                                  self.settings['Name3'])
            self.filedetailedinfo = open_file(self.settings['Name3'], mode='w')

        # if self.settings['Directory2'] is not None:
        #     self.settings['Name2'] = os.path.join(self.settings['Directory2'],
        #                                           self.settings['Name2'])
        #     self.file2 = open_file(self.settings['Name2'], mode='a')

    '''pytables auxiliary'''
    class PyeneHDF5Flags:
        def __init__(self):
            self.settings = {}
            self.settings['treefile'] = True
            self.settings['networkfile'] = True
            self.settings['time'] = True
            self.settings['scenarios'] = True
            self.settings['NoHydro'] = True
            self.settings['NoPump'] = True
            self.settings['NoRES'] = True

            self.devices = {}
            self.results = {}

    class PyeneHDF5Settings(IsDescription):
        '''Simulation settings'''
        treefile = StringCol(itemsize=30, dflt=" ", pos=0)  # File name (tree)
        networkfile = StringCol(itemsize=30, dflt=" ", pos=1)  # FIle (network)
        time = Int16Col(dflt=1, pos=2)  # generatop
        scenarios = Int16Col(dflt=1, pos=3)  # scenarios
        NoHydro = Int16Col(dflt=1, pos=4)  # hydropower plants
        NoPump = Int16Col(dflt=1, pos=5)  # pumps
        NoRES = Int16Col(dflt=1, pos=6)  # RES generators

    class PyeneHDF5Devices(IsDescription):
        '''Characteristics of devices '''
        location = Int16Col(dflt=1, pos=0)  # Location
        max = Float32Col(dflt=1, pos=1)  # Capacity
        cost = Float32Col(dflt=4, pos=2)  # Cost/value
        link = Int16Col(dflt=1, pos=3)  # Location

    class PyeneHDF5Results(IsDescription):
        time = Int16Col(dflt=1, pos=0)  # time period
        generation = Float32Col(dflt=1, pos=1)  # conventional generator
        hydropower = Float32Col(dflt=1, pos=2)  # hydropower
        RES = Float32Col(dflt=1, pos=3)  # RES generator
        demand = Float32Col(dflt=1, pos=4)  # total demand
        pump = Float32Col(dflt=1, pos=5)  # use of pumps
        loss = Float32Col(dflt=1, pos=6)  # losses
        curtailment = Float32Col(dflt=1, pos=7)  # sCurtailment
        spill = Float32Col(dflt=1, pos=8)  # spilling
    
    def initialise(self, dir=None, name_file=None):
        self.fileh = open_file(os.path.join(dir, name_file+".h5"), mode='w')
    
    def save_results(self, model=None, sim_no=None):
        from .pyene_Models import models
        if isinstance(model, models):
            values, starts, characteristics = model.get_outputs()
        HDF5group = \
            self.fileh.create_group(self.fileh.root, 'Simulation_{:05d}'.format(sim_no))
        
        # Extracting name to create subgroups
        information_name = []
        information_problem = []
        information_position_tree = []
        for out in model.data.outputs:
            if out.exist("data device") and out.get_characteristic("data device") == ["h5"]:
                information_name.append(out.get_characteristic("name")[0])
                information_problem.append(out.get_characteristic("problem")[0])
                #TODO: Generalise subscripts
                subscripts = model.get_latest_subscripts()
                for key, value in subscripts.items():
                    if key == "pt":
                        information_position_tree.append(value[0])
                # if out.exist("pt"):
                #     information_position_tree.append(out.get_characteristic("pt"))
                # else:
                #     information_position_tree.append(None)

        used_values = [False for _ in range(len(values))]
        for i_n, i_p, i_pt in zip(information_name, information_problem, information_position_tree):
            if (i_p == "DC OPF"):
                table_columns = {}
                aux = 0
                for cha_group in characteristics:
                    is_group = False
                    for cha in cha_group:
                        if cha[0].decode('utf-8') == "name" and cha[2].decode('utf-8') == i_n:
                            is_group = True
                    if is_group:
                        for cha in cha_group:
                            if cha[0].decode('utf-8') == "pt" or cha[0].decode('utf-8') == "hour":
                                table_columns[cha[0].decode('utf-8')] = StringCol(itemsize=64, dflt='', pos=aux)
                                aux = aux + 1
                        break
                # Adding IDs
                for cha_group in characteristics:
                    is_group = False
                    for cha in cha_group:
                        if cha[0].decode('utf-8') == "name" and cha[2].decode('utf-8') == i_n:
                            is_group = True
                    if is_group:
                        for cha in cha_group:
                            if cha[0].decode('utf-8') == "ID" and cha[2].decode('utf-8') not in table_columns:
                                table_columns[cha[2].decode('utf-8')] = Float32Col(dflt=0, pos=aux)
                                aux = aux + 1
                                break
                if (i_pt is not None):
                    word = "_" + i_pt[0]
                    for c in i_pt[1:]:
                        word = word + "_" + c
                else:
                    word = ""
                HDF5table = \
                self.fileh.create_table(HDF5group, "{}{}".format(i_n, word), table_columns)
                HDF5row = HDF5table.row

                is_end = False
                while not is_end:
                    is_end = True
                    reference = {}
                    for num, (val, u_val, cha_group) in enumerate(zip(values, used_values, characteristics)):
                        is_group = True
                        if not u_val:
                            for cha in cha_group:
                                if (cha[0].decode('utf-8') == "name" and cha[2].decode('utf-8') != i_n):
                                    is_group = False
                                if (cha[0].decode('utf-8') == "pt"):
                                    if (len(cha[2:]) != len(i_pt)): is_group = False
                                    else:
                                        for num1, c in enumerate(cha[2:]):
                                            if (c.decode('utf-8') != i_pt[num1]):
                                                is_group = False
                                                break
                        if is_group and not u_val and not reference:
                            is_end = False
                            for cha in cha_group:
                                if (cha[0].decode('utf-8') == "pt" or cha[0].decode('utf-8') == "hour") and len(cha) == 3:
                                    reference[cha[0].decode('utf-8')] = cha[2].decode('utf-8')
                                    HDF5row[cha[0].decode('utf-8')] = cha[2].decode('utf-8')
                                elif (cha[0].decode('utf-8') == "pt" or cha[0].decode('utf-8') == "hour") and len(cha) > 3:
                                    word = cha[2].decode('utf-8')
                                    for c in cha[3:]:
                                        word = word + "_" + c.decode('utf-8')
                                    reference[cha[0].decode('utf-8')] = word
                                    HDF5row[cha[0].decode('utf-8')] = word
                                elif cha[0].decode('utf-8') == "ID":
                                    HDF5row[cha[2].decode('utf-8')] = val
                                    used_values[num] = True
                        elif is_group and not u_val and reference:
                            compare_cha = {}
                            ID = ""
                            for cha in cha_group:
                                if (cha[0].decode('utf-8') == "pt" or cha[0].decode('utf-8') == "hour") and len(cha) == 3:
                                    compare_cha[cha[0].decode('utf-8')] = cha[2].decode('utf-8')
                                elif (cha[0].decode('utf-8') == "pt" or cha[0].decode('utf-8') == "hour") and len(cha) > 3:
                                    word = cha[2].decode('utf-8')
                                    for c in cha[3:]:
                                        word = word + "_" + c.decode('utf-8')
                                    compare_cha[cha[0].decode('utf-8')] = word
                                elif cha[0].decode('utf-8') == "ID":
                                    ID = cha[2].decode('utf-8')
                            if compare_cha == reference:
                                is_end = False
                                HDF5row[ID] = val
                                used_values[num] = True
                    if not is_end:
                        HDF5row.append()
                HDF5table.flush()
            if (i_p == "BT"):
                table_columns = {}
                aux = 0
                for cha_group in characteristics:
                    is_group = False
                    for cha in cha_group:
                        if cha[0].decode('utf-8') == "name" and cha[2].decode('utf-8') == i_n:
                            is_group = True
                    if is_group:
                        for cha in cha_group:
                            if cha[0].decode('utf-8') != "name" and cha[0].decode('utf-8') != "reference":
                                table_columns[cha[0].decode('utf-8')] = StringCol(itemsize=64, dflt='', pos=aux)
                                aux = aux + 1
                # Adding references
                for cha_group in characteristics:
                    is_group = False
                    for cha in cha_group:
                        if cha[0].decode('utf-8') == "name" and cha[2].decode('utf-8') == i_n:
                            is_group = True
                    if is_group:
                        for cha in cha_group:
                            if cha[0].decode('utf-8') == "reference" and cha[2].decode('utf-8') not in table_columns:
                                table_columns[cha[2].decode('utf-8')] = Float32Col(dflt=0, pos=aux)
                                aux = aux + 1
                                break
                
                HDF5table = \
                self.fileh.create_table(HDF5group, "{}".format(i_n), table_columns)
                HDF5row = HDF5table.row

                is_end = False
                while not is_end:
                    is_end = True
                    reference = {}
                    if (False not in used_values): break
                    for num, (val, u_val, cha_group) in enumerate(zip(values, used_values, characteristics)):
                        is_group = False
                        if not u_val:
                            for cha in cha_group:
                                if cha[0].decode('utf-8') == "name" and cha[2].decode('utf-8') == i_n:
                                    is_group = True
                        if is_group and not u_val and not reference:
                            is_end = False
                            for cha in cha_group:
                                if cha[0].decode('utf-8') != "name" and cha[0].decode('utf-8') != "reference" and len(cha) == 3:
                                    reference[cha[0].decode('utf-8')] = cha[2].decode('utf-8')
                                    HDF5row[cha[0].decode('utf-8')] = cha[2].decode('utf-8')
                                elif cha[0].decode('utf-8') != "name" and cha[0].decode('utf-8') != "reference" and len(cha) > 3:
                                    word = cha[2].decode('utf-8')
                                    for c in cha[3:]:
                                        word = word + "_" + c.decode('utf-8')
                                    reference[cha[0].decode('utf-8')] = word
                                    HDF5row[cha[0].decode('utf-8')] = word
                                elif cha[0].decode('utf-8') == "reference":
                                    HDF5row[cha[2].decode('utf-8')] = val
                                    used_values[num] = True
                        elif is_group and not u_val and reference:
                            compare_cha = {}
                            ID = ""
                            for cha in cha_group:
                                if cha[0].decode('utf-8') != "name" and cha[0].decode('utf-8') != "reference" and len(cha) == 3:
                                    compare_cha[cha[0].decode('utf-8')] = cha[2].decode('utf-8')
                                elif cha[0].decode('utf-8') != "name" and cha[0].decode('utf-8') != "reference" and len(cha) > 3:
                                    word = cha[2].decode('utf-8')
                                    for c in cha[3:]:
                                        word = word + "_" + c.decode('utf-8')
                                    compare_cha[cha[0].decode('utf-8')] = word
                                elif cha[0].decode('utf-8') == "reference":
                                    ID = cha[2].decode('utf-8')
                            if compare_cha == reference:
                                is_end = False
                                HDF5row[ID] = val
                                used_values[num] = True
                    if not is_end:
                        HDF5row.append()
                HDF5table.flush()

    def Accumulate(self, EN, m):
        ''' Accumulate results '''
        if self.settings['Directory2'] is None:
            return

        for x in range(5):
            aux = [False for x in range(5)]
            aux[x] = True
            self.data['cost'][x] += EN.get_OFparts(m, aux)

        self.data['curtailment'] += EN.get_AllDemandCurtailment(m)[0]
        self.data['spill'] += EN.get_AllRES(m)
        self.data['OF'] += EN.get_OFparts(m, [True for x in range(5)])

    def SaveSettings(self, EN):
        ''' Save simulation settings '''
        if self.settings['Directory1'] is None:
            return

        HDF5group = self.fileh.create_group(self.fileh.root, 'Core')
        HDF5table = self.fileh.create_table(HDF5group, "table",
                                            self.PyeneHDF5Settings)
        HDF5row = HDF5table.row
        HDF5row['treefile'] = EN.EM.settings['File']
        HDF5row['networkfile'] = EN.NM.settings['File']
        HDF5row['time'] = EN.NM.settings['NoTime']
        HDF5row['scenarios'] = EN.NM.scenarios['Number']
        HDF5row['NoHydro'] = EN.NM.hydropower['Number']
        HDF5row['NoPump'] = EN.NM.pumps['Number']
        HDF5row['NoRES'] = EN.NM.RES['Number']
        HDF5row.append()
        HDF5table.flush()

        HDF5table = self.fileh.create_table(HDF5group, "Hydpopower_plants",
                                            self.PyeneHDF5Devices)
        HDF5row = HDF5table.row
        for x in EN.NM.Gen.Hydro:
            HDF5row['location'] = x.get_BusPos()
            HDF5row['max'] = x.get_Max()
            HDF5row['cost'] = x.get_Cost()
            HDF5row.append()
        HDF5table.flush()

        HDF5table = self.fileh.create_table(HDF5group, "Pumps",
                                            self.PyeneHDF5Devices)
        HDF5row = HDF5table.row
        for xh in range(EN.NM.pumps['Number']):
            HDF5row['location'] = EN.NM.pumps['Bus'][xh]
            HDF5row['max'] = EN.NM.pumps['Max'][xh]
            HDF5row['cost'] = EN.NM.pumps['Value'][xh]
            HDF5row.append()
        HDF5table.flush()

        HDF5table = self.fileh.create_table(HDF5group, "RES_generators",
                                            self.PyeneHDF5Devices)
        HDF5row = HDF5table.row
        for x in EN.NM.Gen.RES:
            HDF5row['location'] = x.get_BusPos()
            HDF5row['max'] = x.get_Max()
            HDF5row['cost'] = x.get_Cost()
            HDF5row.append()

        HDF5table.flush()

    def saveSummary(self, simulation_name):
        # Independent files
        if self.settings['Directory1'] is not None:
            aux = os.path.join(self.settings['Directory2'], simulation_name)
            fileh = open_file(aux, mode='w')
            fileh.create_array(fileh.root, "OF", self.data['OF'])
            fileh.create_array(fileh.root, "curtailment",
                               self.data['curtailment'])
            fileh.create_array(fileh.root, "spill", self.data['spill'])
            fileh.create_array(fileh.root, "Cost_Component", self.data['cost'])
            fileh.create_array(fileh.root, "time", [self.time['All'],
                                                    self.time['glpk']])
            fileh.close()

    def saveResults(self, EN, model, SimNo):
        if isinstance(model, ConcreteModel):
            self.saveResultspyomo( EN, model, SimNo)
        elif isinstance(model, EnergyandNetwork) or isinstance(model, Networkmodel):
            self.saveResultsGLPK(EN, model, SimNo)

    def saveResultspyomo(self, EN, m, SimNo):
        ''' Save results of each iteration '''

        # Accumulate data
        self.Accumulate(EN, m)

        if self.settings['Directory1'] is None:
            return

        HDF5group = \
            self.fileh.create_group(self.fileh.root,
                                    'Simulation_{:05d}'.format(SimNo))
        HDF5aux = np.zeros((EN.NM.scenarios['NoDem'],
                            EN.NM.settings['NoTime']), dtype=float)

        xp = 0
        for xs in range(EN.NM.scenarios['NoDem']):
            for xt in range(EN.NM.settings['NoTime']):
                HDF5aux[xs][xt] = EN.NM.scenarios['Demand'][xp]
                xp += 1

        self.fileh.create_array(HDF5group, "Demand_profiles", HDF5aux)

        HDF5aux = np.zeros((EN.NM.scenarios['NoRES'],
                            EN.NM.settings['NoTime']), dtype=float)
        xp = 0
        for xs in range(EN.NM.scenarios['NoRES']):
            for xt in range(EN.NM.settings['NoTime']):
                HDF5aux[xs][xt] = EN.NM.scenarios['RES'][xp] * \
                    EN.NM.ENetwork.get_Base()
                xp += 1
        self.fileh.create_array(HDF5group, "RES_profiles", HDF5aux)

        # Hydropower allowance
        aux = np.zeros(EN.EM.settings['Vectors'], dtype=float)
        if type(m.vEIn) is np.ndarray:
            if EN.EM.settings['Vectors'] == 1:
                aux[0] = EN.EM.Weight['In'][1]
            else:
                for xv in EN.EM.s['Vec']:
                    aux[xv] = m.vEIn[1][xv]
        else:
            for xv in EN.EM.s['Vec']:
                aux[xv] = m.vEIn[1, xv].value

        self.fileh.create_array(HDF5group, "Hydro_Allowance", aux)

        hp_marginal = np.zeros(EN.EM.settings['Vectors'], dtype=float)
        for xi in range(EN.EM.settings['Vectors']):
            hp_marginal[xi] = EN.get_HydroMarginal(m, xi+1)
        self.fileh.create_array(HDF5group, "Hydro_Marginal", hp_marginal)

        for xs in range(EN.NM.scenarios['Number']):
            HDF5table = \
                self.fileh.create_table(HDF5group,
                                        "Scenario_{:02d}".format(xs),
                                        self.PyeneHDF5Results)
            HDF5row = HDF5table.row
            for xt in range(EN.NM.settings['NoTime']):
                HDF5row['time'] = xt
                HDF5row['generation'] = \
                    EN.get_AllGeneration(m, 'Conv', 'snapshot', times=[xt],
                                         scens=[xs])
                HDF5row['hydropower'] = \
                    EN.get_AllGeneration(m, 'Hydro', 'snapshot', times=[xt],
                                         scens=[xs])
                HDF5row['RES'] = EN.get_AllGeneration(m, 'RES', 'snapshot',
                                                      times=[xt], scens=[xs])
                HDF5row['spill'] = EN.get_AllRES(m, 'snapshot', times=[xt],
                                                 scens=[xs])
                HDF5row['demand'] = EN.get_AllDemand(m, 'snapshot',
                                                     times=[xt], scens=[xs])[0]
                HDF5row['pump'] = EN.get_AllPumps(m, 'snapshot', times=[xt],
                                                  scens=[xs])
                HDF5row['loss'] = EN.get_AllLoss(m, 'snapshot', times=[xt],
                                                 scens=[xs])
                HDF5row['curtailment'] = \
                    EN.get_AllDemandCurtailment(m, 'snapshot', times=[xt],
                                                scens=[xs])[0]
                HDF5row.append()
            HDF5table.flush()

    def saveResultsGLPK(self, EN, GLPKobj, SimNo):
        ''' Save results of each iteration '''

        # Accumulate data
        # self.Accumulate(EN, m)

        if self.settings['Directory1'] is None:
            return

        HDF5group = \
            self.fileh.create_group(self.fileh.root,
                                    'Simulation_{:05d}'.format(SimNo))
        HDF5aux = np.zeros((GLPKobj.NumberDemScenarios,
                            GLPKobj.ShortTemporalConnections), dtype=float)

        xp = 0
        for xs in range(GLPKobj.NumberDemScenarios):
            for xt in range(GLPKobj.ShortTemporalConnections):
                HDF5aux[xs][xt] = EN.NM.scenarios['Demand'][xp]
                xp += 1

        self.fileh.create_array(HDF5group, "Demand_profiles", HDF5aux)

        HDF5aux = np.zeros((EN.NM.scenarios['NoRES'],
                            EN.NM.settings['NoTime']), dtype=float)
        xp = 0
        for xs in range(EN.NM.scenarios['NoRES']):
            for xt in range(EN.NM.settings['NoTime']):
                HDF5aux[xs][xt] = EN.NM.scenarios['RES'][xp] * \
                    EN.NM.ENetwork.get_Base()
                xp += 1
        self.fileh.create_array(HDF5group, "RES_profiles", HDF5aux)

        # Hydropower allowance
        aux = np.zeros(EN.EM.settings['Vectors'], dtype=float)
        if EN.EM.settings['Vectors'] == 1:
            aux[0] = EN.EM.Weight['In'][1]
        else:
            for xv in range(EN.EM.settings['Vectors']):
                aux[xv] = GLPKobj.IntakeTree[1, xv]

        self.fileh.create_array(HDF5group, "Hydro_Allowance", aux)

        hp_marginal = np.zeros(EN.EM.settings['Vectors'], dtype=float)
        for xi in range(EN.EM.settings['Vectors']):
            hp_marginal[xi] = 0.0
        self.fileh.create_array(HDF5group, "Hydro_Marginal", hp_marginal)

        # Getting the solution from GLPK

        ThermalGeneration = GLPKobj.GetThermalGeneration()
        RESGeneration = GLPKobj.GetRESGeneration()
        HydroGeneration = GLPKobj.GetHydroGeneration()
        PumpOperation = GLPKobj.GetPumpOperation()
        if GLPKobj.FlagProblem and GLPKobj.FlagFeasibility:
            LoadCurtailment = GLPKobj.GetLoadCurtailmentNodes()
        else:
            LoadCurtailment = GLPKobj.GetLoadCurtailmentSystemED()
        
        GenerationCurtailment = GLPKobj.GetGenerationCurtailmentNodes()
        ActivePowerFlow = GLPKobj.GetActivePowerFlow()
        Branches = EN.NM.ENetwork.Branch
        if GLPKobj.FlagProblem and GLPKobj.LossesFlag:
            ActivePowerLosses = GLPKobj.GetActivePowerLosses()
        elif not GLPKobj.LossesFlag and GLPKobj.PercentageLosses is not None and \
            GLPKobj.FlagProblem:
            # Interpolation of losses
            ActivePowerLosses = \
                np.empty((len(GLPKobj.LongTemporalConnections),\
                    GLPKobj.ShortTemporalConnections, \
                    (GLPKobj.NumberContingencies + 1), \
                    GLPKobj.NumberLinesPS))
            for xh in GLPKobj.LongTemporalConnections:
                for xt in range(GLPKobj.ShortTemporalConnections):
                    FullLoss = 0
                    # Add all power generation
                    if GLPKobj.NumberConvGen > 0:
                        for xn in range(GLPKobj.NumberConvGen):
                            FullLoss += ThermalGeneration[xh, xt, xn]
                    if GLPKobj.NumberRESGen > 0:
                        for xn in range(GLPKobj.NumberRESGen):
                            FullLoss += RESGeneration[xh, xt, xn]
                    if GLPKobj.NumberHydroGen > 0:
                        for xn in range(GLPKobj.NumberHydroGen):
                            FullLoss += HydroGeneration[xh, xt, xn]
                                        
                    # Substract all power generation curtailment
                    if GLPKobj.NumberConvGen > 0:
                        for xn in range(GLPKobj.NumberConvGen):
                            for xco in range(GLPKobj.NumberContingencies + 1):
                                FullLoss -= ThermalGenerationCurtailment\
                                    [xh, xt, xco, xn]
                    if GLPKobj.NumberRESGen > 0:
                        for xn in range(GLPKobj.NumberRESGen):
                            for xco in range(GLPKobj.NumberContingencies + 1):
                                FullLoss -= RESGenerationCurtailment\
                                    [xh, xt, xco, xn]
                    if GLPKobj.NumberHydroGen > 0:
                        for xn in range(GLPKobj.NumberHydroGen):
                            for xco in range(GLPKobj.NumberContingencies + 1):
                                FullLoss -= HydroGenerationCurtailment\
                                    [xh, xt, xco, xn]

                    # Substract demand
                    for xn in range(GLPKobj.NumberNodesPS):
                        if GLPKobj.NumberDemScenarios == 0:
                            FullLoss -= GLPKobj.PowerDemandNode[xn] * \
                                GLPKobj.MultScenariosDemand[xh, xn] * \
                                GLPKobj.BaseUnitPower
                        else:
                            FullLoss -= GLPKobj.PowerDemandNode[xn] * \
                                GLPKobj.MultScenariosDemand[xh, xt, xn] * \
                                GLPKobj.BaseUnitPower

                        # Curtailment
                        if GLPKobj.FlagFeasibility:
                            # Add load curtailment
                            for xco in range(GLPKobj.NumberContingencies + 1):
                                FullLoss += LoadCurtailment[xh, xt, xco, xn]

                    # Substract non-technical losses
                    for xb in range(GLPKobj.NumberLinesPS):
                        FullLoss -= Branches[xb].getLoss()

                    # Allocate losses per line
                    FullFlow = 0
                    for xb in range(GLPKobj.NumberLinesPS):
                        for xco in range(GLPKobj.NumberContingencies + 1):
                            FullFlow += abs(ActivePowerFlow[xh, xt, xco, xb])
                    if FullFlow > 0:
                        for xb in range(GLPKobj.NumberLinesPS):
                            for xco in range(GLPKobj.NumberContingencies + 1):
                                aux = abs(ActivePowerFlow[xh, xt, xco, xb]) / FullFlow
                            ActivePowerLosses[xh, xt, xco, xb] = FullLoss * aux + \
                                Branches[xb].getLoss()
        else:
            ActivePowerLosses = \
                np.zeros((len(GLPKobj.LongTemporalConnections),\
                    GLPKobj.ShortTemporalConnections, \
                    (GLPKobj.NumberContingencies + 1), \
                    GLPKobj.NumberLinesPS))

        data_descr = dict(
            time = Int16Col(dflt=1, pos=0),  # time period
            generation = Float32Col(dflt=1, pos=1),  # conventional generator
            hydropower = Float32Col(dflt=1, pos=2),  # hydropower
            RES = Float32Col(dflt=1, pos=3),  # RES generator
            demand = Float32Col(dflt=1, pos=4),  # total demand
            pump = Float32Col(dflt=1, pos=5),  # use of pumps
            loss = Float32Col(dflt=1, pos=6),  # losses
            curtailment = Float32Col(dflt=1, pos=7),  # sCurtailment
            spill = Float32Col(dflt=1, pos=8),  # spilling
            gen_cur = Float32Col(dflt=1, pos=9),  # spilling
        )

        counter = 10

        if RESGeneration is not None:
            for k in range(GLPKobj.NumberRESGen):
                data_descr['RES_Gen_{}_Pot_{}'.format(k, GLPKobj.MaxRESGen[k])] = \
                    Float32Col(dflt=1, pos=counter+k)
            counter = counter + GLPKobj.NumberRESGen
        
        if HydroGeneration is not None:
            for k in range(GLPKobj.NumberHydroGen):
                data_descr['Hydro_Gen_{}_Pot_{}'.format(k, GLPKobj.MaxHydroGen[k])] = \
                    Float32Col(dflt=1, pos=counter+k)
        

        for xs in GLPKobj.LongTemporalConnections:
            HDF5table = \
                self.fileh.create_table(HDF5group,
                                        "Scenario_{:02d}".format(xs),
                                        data_descr)
            HDF5row = HDF5table.row
            for xt in range(GLPKobj.ShortTemporalConnections):
                HDF5row['time'] = xt
                auxvar = 0
                if ThermalGeneration is not None:
                    for k in range(GLPKobj.NumberConvGen):
                        auxvar += ThermalGeneration[xs, xt, k]
                HDF5row['generation'] = auxvar

                
                auxvar = 0
                if HydroGeneration is not None:
                    for k in range(GLPKobj.NumberHydroGen):
                        auxvar += HydroGeneration[xs, xt, k]
                HDF5row['hydropower'] = auxvar
                
                auxvar = 0
                if RESGeneration is not None:
                    for k in range(GLPKobj.NumberRESGen):
                        auxvar += RESGeneration[xs, xt, k]
                HDF5row['RES'] = auxvar

                HDF5row['spill'] = 0

                if GenerationCurtailment is not None:
                    auxvar = 0
                    for xco in range(GLPKobj.NumberContingencies + 1):
                        for xnod in range(EN.NM.ENetwork.get_NoBus()):
                            auxvar += GenerationCurtailment[\
                                xs, xt, xco, xnod]
                HDF5row['gen_cur'] = auxvar

                totaldemand = 0 
                for k in range(EN.NM.ENetwork.get_NoBus()):
                    # TODO: Change the inputs of losses and demand scenarios
                    # for parameters
                    if GLPKobj.TypeNode[k] != 4:
                        if GLPKobj.NumberDemScenarios == 0:
                            totaldemand = totaldemand + \
                                GLPKobj.PowerDemandNode[k] * \
                                GLPKobj.MultScenariosDemand[xs, k] * \
                                    GLPKobj.BaseUnitPower
                        else:
                            totaldemand = totaldemand + \
                                GLPKobj.PowerDemandNode[k] * \
                                GLPKobj.MultScenariosDemand[xs, xt, k] * \
                                    GLPKobj.BaseUnitPower
                HDF5row['demand'] = totaldemand                

                auxvar = 0
                if PumpOperation is not None:
                    for k in range(GLPKobj.NumberPumps):
                        auxvar += PumpOperation[xs, xt, k]
                HDF5row['pump'] = auxvar

                auxvar = 0
                if EN.NM.settings['Losses'] or GLPKobj.PercentageLosses is not None:
                    for k in range(GLPKobj.NumberContingencies + 1):
                        for ii in range(GLPKobj.NumberLinesPS):
                            auxvar += ActivePowerLosses[xs, xt, k, ii]
                HDF5row['loss'] = auxvar

                auxvar = 0
                if GLPKobj.FlagProblem and LoadCurtailment is not None:
                    for k in range(GLPKobj.NumberContingencies + 1):
                        for ii in range(GLPKobj.NumberNodesPS):
                            auxvar += LoadCurtailment[xs, xt, k, ii]
                if not GLPKobj.FlagProblem and LoadCurtailment is not None:
                    auxvar += LoadCurtailment[xs, xt]
                HDF5row['curtailment'] = auxvar

                if RESGeneration is not None:
                    for k in range(GLPKobj.NumberRESGen):
                        HDF5row['RES_Gen_{}_Pot_{}'.format(k, \
                        GLPKobj.MaxRESGen[k])] = \
                            RESGeneration[xs, xt, k]
                
                if HydroGeneration is not None:
                    for k in range(GLPKobj.NumberHydroGen):
                        HDF5row['Hydro_Gen_{}_Pot_{}'.format(k, \
                        GLPKobj.MaxHydroGen[k])] = \
                            HydroGeneration[xs, xt, k]

                HDF5row.append()
            HDF5table.flush()

    def SaveDetailedResultsGLPK(self, EN, GLPKobj, SimNo):
        ''' Save results of each iteration '''

        # Accumulate data
        # self.Accumulate(EN, m)

        if self.settings['Directory1'] is None:
            return

        HDF5group = \
            self.filedetailedinfo.create_group(self.filedetailedinfo.root,
                                    'Simulation_{:05d}'.format(SimNo))
        HDF5aux = np.zeros((GLPKobj.NumberDemScenarios,
                            GLPKobj.ShortTemporalConnections), dtype=float)

        xp = 0
        for xs in range(GLPKobj.NumberDemScenarios):
            for xt in range(GLPKobj.ShortTemporalConnections):
                HDF5aux[xs][xt] = EN.NM.scenarios['Demand'][xp]
                xp += 1

        self.filedetailedinfo.create_array(HDF5group, "Demand_profiles", HDF5aux)

        HDF5aux = np.zeros((EN.NM.scenarios['NoRES'],
                            EN.NM.settings['NoTime']), dtype=float)
        xp = 0
        for xs in range(EN.NM.scenarios['NoRES']):
            for xt in range(EN.NM.settings['NoTime']):
                HDF5aux[xs][xt] = EN.NM.scenarios['RES'][xp] * \
                    EN.NM.ENetwork.get_Base()
                xp += 1
        self.filedetailedinfo.create_array(HDF5group, "RES_profiles", HDF5aux)

        # Hydropower allowance
        aux = np.zeros(EN.EM.settings['Vectors'], dtype=float)
        if EN.EM.settings['Vectors'] == 1:
            aux[0] = EN.EM.Weight['In'][1]
        else:
            for xv in range(EN.EM.settings['Vectors']):
                aux[xv] = GLPKobj.IntakeTree[1, xv]

        self.filedetailedinfo.create_array(HDF5group, "Hydro_Allowance", aux)

        hp_marginal = np.zeros(EN.EM.settings['Vectors'], dtype=float)
        for xi in range(EN.EM.settings['Vectors']):
            hp_marginal[xi] = 0.0
        self.filedetailedinfo.create_array(HDF5group, "Hydro_Marginal", hp_marginal)

        # Getting the solution from GLPK
        ThermalGeneration = GLPKobj.GetThermalGeneration()
        RESGeneration = GLPKobj.GetRESGeneration()
        HydroGeneration = GLPKobj.GetHydroGeneration()
        PumpOperation = GLPKobj.GetPumpOperation()
        if GLPKobj.FlagProblem and GLPKobj.FlagFeasibility:
            LoadCurtailment = GLPKobj.GetLoadCurtailmentNodes()
        else:
            LoadCurtailment = GLPKobj.GetLoadCurtailmentSystemED()
        VoltageAngle = GLPKobj.GetVoltageAngle()
        ActivePowerFlow = GLPKobj.GetActivePowerFlow()
        ThermalGenerationCurtailment = GLPKobj.GetThermalGenerationCurtailmentNodes()
        RESGenerationCurtailment = GLPKobj.GetRESGenerationCurtailmentNodes()
        HydroGenerationCurtailment = GLPKobj.GetHydroGenerationCurtailmentNodes()
        Branches = EN.NM.ENetwork.Branch
        if GLPKobj.FlagProblem and GLPKobj.LossesFlag:
            ActivePowerLosses = GLPKobj.GetActivePowerLosses()
        elif not GLPKobj.LossesFlag and GLPKobj.PercentageLosses is not None and \
            GLPKobj.FlagProblem:
            # Interpolation of losses
            ActivePowerLosses = \
                np.empty((len(GLPKobj.LongTemporalConnections),\
                    GLPKobj.ShortTemporalConnections, \
                    (GLPKobj.NumberContingencies + 1), \
                    GLPKobj.NumberLinesPS))
            for xh in GLPKobj.LongTemporalConnections:
                for xt in range(GLPKobj.ShortTemporalConnections):
                    FullLoss = 0
                    # Add all power generation
                    if GLPKobj.NumberConvGen > 0:
                        for xn in range(GLPKobj.NumberConvGen):
                            FullLoss += ThermalGeneration[xh, xt, xn]
                    if GLPKobj.NumberRESGen > 0:
                        for xn in range(GLPKobj.NumberRESGen):
                            FullLoss += RESGeneration[xh, xt, xn]
                    if GLPKobj.NumberHydroGen > 0:
                        for xn in range(GLPKobj.NumberHydroGen):
                            FullLoss += HydroGeneration[xh, xt, xn]
                                        
                    # Substract all power generation curtailment
                    if GLPKobj.NumberConvGen > 0:
                        for xn in range(GLPKobj.NumberConvGen):
                            for xco in range(GLPKobj.NumberContingencies + 1):
                                FullLoss -= ThermalGenerationCurtailment\
                                    [xh, xt, xco, xn]
                    if GLPKobj.NumberRESGen > 0:
                        for xn in range(GLPKobj.NumberRESGen):
                            for xco in range(GLPKobj.NumberContingencies + 1):
                                FullLoss -= RESGenerationCurtailment\
                                    [xh, xt, xco, xn]
                    if GLPKobj.NumberHydroGen > 0:
                        for xn in range(GLPKobj.NumberHydroGen):
                            for xco in range(GLPKobj.NumberContingencies + 1):
                                FullLoss -= HydroGenerationCurtailment\
                                    [xh, xt, xco, xn]

                    # Substract demand
                    for xn in range(GLPKobj.NumberNodesPS):
                        if GLPKobj.NumberDemScenarios == 0:
                            FullLoss -= GLPKobj.PowerDemandNode[xn] * \
                                GLPKobj.MultScenariosDemand[xh, xn] * \
                                GLPKobj.BaseUnitPower
                        else:
                            FullLoss -= GLPKobj.PowerDemandNode[xn] * \
                                GLPKobj.MultScenariosDemand[xh, xt, xn] * \
                                GLPKobj.BaseUnitPower

                        # Curtailment
                        if GLPKobj.FlagFeasibility:
                            # Add load curtailment
                            for xco in range(GLPKobj.NumberContingencies + 1):
                                FullLoss += LoadCurtailment[xh, xt, xco, xn]

                    # Substract non-technical losses
                    for xb in range(GLPKobj.NumberLinesPS):
                        FullLoss -= Branches[xb].getLoss()

                    # Allocate losses per line
                    FullFlow = 0
                    for xb in range(GLPKobj.NumberLinesPS):
                        for xco in range(GLPKobj.NumberContingencies + 1):
                            FullFlow += abs(ActivePowerFlow[xh, xt, xco, xb])
                    if FullFlow > 0:
                        for xb in range(GLPKobj.NumberLinesPS):
                            for xco in range(GLPKobj.NumberContingencies + 1):
                                aux = abs(ActivePowerFlow[xh, xt, xco, xb]) / FullFlow
                            ActivePowerLosses[xh, xt, xco, xb] = FullLoss * aux + \
                                Branches[xb].getLoss()
        else:
            ActivePowerLosses = \
                np.zeros((len(GLPKobj.LongTemporalConnections),\
                    GLPKobj.ShortTemporalConnections, \
                    (GLPKobj.NumberContingencies + 1), \
                    GLPKobj.NumberLinesPS))
            
        data_descr_RES = dict(
            time = Int16Col(dflt=1, pos=0)  # time period
        )

        for k in range(GLPKobj.NumberRESGen):
            data_descr_RES[str(GLPKobj.OriginalNumberRESGen[k])] = \
                Float32Col(dflt=1, pos=1+k)
        
        for xs in GLPKobj.LongTemporalConnections:
            HDF5table = \
                self.fileh.create_table(HDF5group,
                                        "Scenario_{:02d}_RES".format(xs),
                                        data_descr_RES)
            HDF5row = HDF5table.row
            for xt in range(GLPKobj.ShortTemporalConnections):
                HDF5row['time'] = xt
                if RESGeneration is not None:
                    for k in range(GLPKobj.NumberRESGen):
                        HDF5row[str(GLPKobj.OriginalNumberRESGen[k])] = \
                            RESGeneration[xs, xt, k]
                HDF5row.append()
            HDF5table.flush()
        
        data_descr_Hydro = dict(
            time = Int16Col(dflt=1, pos=0)  # time period
        )

        for k in range(GLPKobj.NumberHydroGen):
            data_descr_Hydro[str(GLPKobj.OriginalNumberHydroGen[k])] = \
                Float32Col(dflt=1, pos=1+k)
        
        for xs in GLPKobj.LongTemporalConnections:
            HDF5table = \
                self.fileh.create_table(HDF5group,
                                        "Scenario_{:02d}_Hydro".format(xs),
                                        data_descr_Hydro)
            HDF5row = HDF5table.row
            for xt in range(GLPKobj.ShortTemporalConnections):
                HDF5row['time'] = xt
                if HydroGeneration is not None:
                    for k in range(GLPKobj.NumberHydroGen):
                        HDF5row[str(GLPKobj.OriginalNumberHydroGen[k])] = \
                            HydroGeneration[xs, xt, k]
                HDF5row.append()
            HDF5table.flush()

        # if ThermalGeneration is not None:
        #     self.filedetailedinfo.create_array(HDF5group, "Thermal_Generation", \
        #         ThermalGeneration)
        
        # if RESGeneration is not None:
        #     self.filedetailedinfo.create_array(HDF5group, "RES_Generation", \
        #         RESGeneration)
        
        # if HydroGeneration is not None:
        #     self.filedetailedinfo.create_array(HDF5group, "Hydro_Generation", \
        #         HydroGeneration)
        
        # if PumpOperation is not None:
        #     self.filedetailedinfo.create_array(HDF5group, "Pump_Operation", \
        #         PumpOperation)
        
        # if LoadCurtailment is not None:
        #     self.filedetailedinfo.create_array(HDF5group, "Load_Curtailment", \
        #         LoadCurtailment)
        
        # if ActivePowerLosses is not None:
        #     self.filedetailedinfo.create_array(HDF5group, "Active_Power_Losses", \
        #         ActivePowerLosses)
        
        # if VoltageAngle is not None:
        #     self.filedetailedinfo.create_array(HDF5group, "Voltage_Angle", \
        #         VoltageAngle)
        
        # if ActivePowerFlow is not None:
        #     self.filedetailedinfo.create_array(HDF5group, "Active_Power_Flow", \
        #         ActivePowerFlow)

    def terminate(self, new_implementation=False):

        if not new_implementation:
            if self.settings['Directory1'] is None:
                return
            self.fileh.close()
            self.filedetailedinfo.close()
        else:
            self.fileh.close()

class PrintinScreen():
    '''This class contains all definitions that allows printing on the screen any 
    information requiested by the users'''
    def __init__(self, obj=None):
        if obj is None:
            assert('No information has been passed to this class')
        # Copy attributes
        for pars in obj.__dict__.keys():
            setattr(self, pars, getattr(obj, pars))
        self.PrintinScreenOptions = {
                'Generation': True,
                'Flows': True,
                'Voltages': True,
                'Losses': True,
                'Curtailment': True,
                'Feasibility': True,
                'Services': True,
                'GenBus': True,
                'UC': True,
                'SoC': True,
                'OF': True
                }
        self.lossesresultsOPF = None
    def printlosses(self):
        if self.PrintinScreenOptions['Losses'] and self.settings['Flag']:
            for xh in sh:
                print("\nEPower_Loss=[")
                LossDt = self.printLosses(m, xh)
                for xb in range(self.ENetwork.get_NoBra()):
                    for xt in self.s['Tim']:
                        print("%8.4f " % LossDt[xb][xt], end='')
                    print()
                print("];")

    def printallNetworkResult(self, obj=None):
        ''' This class method prints on the screen all results for the \
            network model'''
        if obj is None:
            print('No object with GLPK results has been passed - Aborting \
                printing of results')
            return

        Generator = self.NM.Gen
        Branches = self.NM.ENetwork.Branch
        ThermalGeneration = obj.GetThermalGeneration()
        RESGeneration = obj.GetRESGeneration()
        HydroGeneration = obj.GetHydroGeneration()
        PumpOperation = obj.GetPumpOperation()
        if not obj.FlagProblem and obj.FlagFeasibility:
            LoadCurtailment = obj.GetLoadCurtailmentSystemED()
        elif obj.FlagProblem and obj.FlagFeasibility:
            LoadCurtailment = obj.GetLoadCurtailmentNodes()
        GenerationCurtailment = obj.GetGenerationCurtailmentNodes()
        VoltageAngle = obj.GetVoltageAngle()
        ActivePowerFlow = obj.GetActivePowerFlow()

        if obj.FlagProblem and obj.LossesFlag:
            ActivePowerLosses = obj.GetActivePowerLosses()
        elif not obj.LossesFlag and obj.PercentageLosses is not None and \
            obj.FlagProblem:
            # Interpolation of losses
            ActivePowerLosses = \
                np.empty((len(obj.LongTemporalConnections),\
                    obj.ShortTemporalConnections, \
                    (obj.NumberContingencies + 1), \
                    obj.NumberLinesPS))
            for xh in obj.LongTemporalConnections:
                for xt in range(obj.ShortTemporalConnections):
                    FullLoss = 0
                    # Add all power generation
                    if obj.NumberConvGen > 0:
                        for xn in range(obj.NumberConvGen):
                            FullLoss += ThermalGeneration[xh, xt, xn]
                    if obj.NumberRESGen > 0:
                        for xn in range(obj.NumberRESGen):
                            FullLoss += RESGeneration[xh, xt, xn]
                    if obj.NumberHydroGen > 0:
                        for xn in range(obj.NumberHydroGen):
                            FullLoss += HydroGeneration[xh, xt, xn]
                                        
                    # Substract all power generation curtailment
                    if obj.NumberConvGen > 0:
                        for xn in range(obj.NumberConvGen):
                            for xco in range(obj.NumberContingencies + 1):
                                FullLoss -= ThermalGenerationCurtailment\
                                    [xh, xt, xco, xn]
                    if obj.NumberRESGen > 0:
                        for xn in range(obj.NumberRESGen):
                            for xco in range(obj.NumberContingencies + 1):
                                FullLoss -= RESGenerationCurtailment\
                                    [xh, xt, xco, xn]
                    if obj.NumberHydroGen > 0:
                        for xn in range(obj.NumberHydroGen):
                            for xco in range(obj.NumberContingencies + 1):
                                FullLoss -= HydroGenerationCurtailment\
                                    [xh, xt, xco, xn]

                    # Substract demand
                    for xn in range(obj.NumberNodesPS):
                        if obj.NumberDemScenarios == 0:
                            FullLoss -= obj.PowerDemandNode[xn] * \
                                obj.MultScenariosDemand[xh, xn] * \
                                obj.BaseUnitPower
                        else:
                            FullLoss -= obj.PowerDemandNode[xn] * \
                                obj.MultScenariosDemand[xh, xt, xn] * \
                                obj.BaseUnitPower

                        # Curtailment
                        if obj.FlagFeasibility:
                            # Add load curtailment
                            for xco in range(obj.NumberContingencies + 1):
                                FullLoss += LoadCurtailment[xh, xt, xco, xn]

                    # Substract non-technical losses
                    for xb in range(obj.NumberLinesPS):
                        FullLoss -= Branches[xb].getLoss()

                    # Allocate losses per line
                    FullFlow = 0
                    for xb in range(obj.NumberLinesPS):
                        for xco in range(obj.NumberContingencies + 1):
                            FullFlow += abs(ActivePowerFlow[xh, xt, xco, xb])
                    if FullFlow > 0:
                        for xb in range(obj.NumberLinesPS):
                            for xco in range(obj.NumberContingencies + 1):
                                aux = abs(ActivePowerFlow[xh, xt, xco, xb]) / FullFlow
                            ActivePowerLosses[xh, xt, xco, xb] = FullLoss * aux + \
                                Branches[xb].getLoss()
        else:
            ActivePowerLosses = \
                np.zeros((len(obj.LongTemporalConnections),\
                    obj.ShortTemporalConnections, \
                    (obj.NumberContingencies + 1), \
                    obj.NumberLinesPS))
        
        # Printing results


        for xh in obj.LongTemporalConnections:
            print("\n% CASE:", xh)

            if self.PrintinScreenOptions['GenBus']:
                print('\nFlow_EGen_Bus=', Generator.get_GenDataAll(), ';')
            
            print("\nDemand=[")
            total_demand = 0
            total_demand_hours = [0 for _ in range(24)]
            for k in range(self.NM.ENetwork.get_NoBus()):
                if obj.TypeNode[k] != 4:
                    for xt in range(obj.ShortTemporalConnections):
                        # TODO: Change the inputs of losses and demand scenarios
                        # for parameters
                        if obj.NumberDemScenarios == 0:
                            val = obj.PowerDemandNode[k] * \
                                obj.MultScenariosDemand[xh, k] * \
                                    obj.BaseUnitPower
                            total_demand += val
                            total_demand_hours[xt] += val
                            print("%8.4f " % val, end='')
                        else:
                            val = obj.PowerDemandNode[k] * \
                                obj.MultScenariosDemand[xh, xt, k] * \
                                    obj.BaseUnitPower
                            print("%8.4f " % val, end='')
                            total_demand += val
                            total_demand_hours[xt] += val
                else:
                    for xt in range(obj.ShortTemporalConnections):
                        print("0.0 ", end='')
                print()
            print("];")

            if self.PrintinScreenOptions['Generation']:
                print("\nFlow_EGen=[")
                total_thermal_gen = 0.0
                total_hydro_gen = 0.0
                total_RES_gen = 0.0
                total_thermal_gen_hours = [0 for _ in range(24)]
                total_hydro_gen_hours = [0 for _ in range(24)]
                total_RES_gen_hours = [0 for _ in range(24)]
                if obj.NumberConvGen > 0:
                    for xn in range(obj.NumberConvGen):
                        for xt in range(obj.ShortTemporalConnections):
                            print("%8.4f " % ThermalGeneration[xh, xt, xn], \
                                end='')
                            total_thermal_gen += ThermalGeneration[xh, xt, xn]
                            total_thermal_gen_hours[xt] += ThermalGeneration[xh, xt, xn]
                        print()
                if obj.NumberRESGen > 0:
                    for xn in range(obj.NumberRESGen):
                        for xt in range(obj.ShortTemporalConnections):
                            print("%8.4f " % RESGeneration[xh, xt, xn], \
                                end='')
                            total_RES_gen += RESGeneration[xh, xt, xn]
                            total_RES_gen_hours[xt] += RESGeneration[xh, xt, xn]
                        print()
                if obj.NumberHydroGen > 0:
                    for xn in range(obj.NumberHydroGen):
                        for xt in range(obj.ShortTemporalConnections):
                            print("%8.4f " % HydroGeneration[xh, xt, xn], \
                                end='')
                            total_hydro_gen += HydroGeneration[xh, xt, xn]
                            total_hydro_gen_hours[xt] += HydroGeneration[xh, xt, xn]
                        print()
                print("];")

            if self.PrintinScreenOptions['UC']:
                print("\nBin_EGen=[")
                aux = 1
                if obj.NumberConvGen > 0:
                    for xn in range(obj.NumberConvGen):
                        for xt in range(obj.ShortTemporalConnections):
                            print("%2.0f " % aux, end='')
                        print()
                if obj.NumberRESGen > 0:
                    for xn in range(obj.NumberRESGen):
                        for xt in range(obj.ShortTemporalConnections):
                            print("%2.0f " % aux, end='')
                        print()
                if obj.NumberHydroGen > 0:
                    for xn in range(obj.NumberHydroGen):
                        for xt in range(obj.ShortTemporalConnections):
                            print("%2.0f " % aux, end='')
                        print()
                print("];")

            if self.PrintinScreenOptions['Flows'] and obj.FlagProblem:
                print("\nFlow_EPower=[")
                for xb in range(obj.NumberLinesPS):
                    for xco in range(obj.NumberContingencies + 1):
                        for xt in range(obj.ShortTemporalConnections):
                            print("%8.4f " % ActivePowerFlow[xh, xt, xco, xb], \
                                end='')
                        print()
                print("];")

            if self.PrintinScreenOptions['Voltages'] and obj.FlagProblem:
                print("\nVoltage_Angle=[")
                for xn in range(obj.NumberNodesPS):
                    for xco in range(obj.NumberContingencies + 1):
                        for xt in range(obj.ShortTemporalConnections):
                            print("%8.4f " % VoltageAngle[xh, xt, xco, xn], \
                                end='')
                        print()
                print("];")
            
            total_losses = 0.0
            if self.PrintinScreenOptions['Losses'] and obj.FlagProblem:
                print("\nEPower_Loss=[")
                for xb in range(obj.NumberLinesPS):
                    for xco in range(obj.NumberContingencies + 1):
                        for xt in range(obj.ShortTemporalConnections):
                            print("%8.4f " % ActivePowerLosses[xh, xt, xco, xb]\
                                , end='')
                            total_losses += ActivePowerLosses[xh, xt, xco, xb]
                    print()
                print("];")

            if self.PrintinScreenOptions['Curtailment']:
                print("\nPumps=[")
                if obj.NumberPumps > 0:
                    for xpu in range(obj.NumberPumps):
                        for xt in range(obj.ShortTemporalConnections):
                            print("%8.4f " % PumpOperation[xh, xt, xpu], end='')
                        print()
                print("];")

            total_load_curtailment = 0.0
            total_load_curtailment_hours = [0 for _ in range(24)]
            if self.PrintinScreenOptions['Feasibility']:
                print("\nFeas=[")
                for xn in range(obj.NumberNodesPS):
                    for xco in range(obj.NumberContingencies + 1):
                        for xt in range(obj.ShortTemporalConnections):
                            if not obj.FlagFeasibility:
                                aux = 0
                            else:
                                aux = LoadCurtailment[xh, xt, xco, xn]
                            print("%8.4f " % aux, end='')
                            total_load_curtailment += aux
                            total_load_curtailment_hours[xt] += aux
                    print()
                print("];")
            print()

            total_gen_cur = 0.0
            total_gen_cur_hours = [0 for _ in range(24)]
            if self.PrintinScreenOptions['Feasibility']:
                print("\nFeasGC=[")
                for xn in range(obj.NumberNodesPS):
                    for xco in range(obj.NumberContingencies + 1):
                        for xt in range(obj.ShortTemporalConnections):
                            if not obj.FlagFeasibility:
                                aux = 0
                            else:
                                aux = GenerationCurtailment\
                                    [xh, xt, xco, xn]
                            print("%8.4f " % aux, end='')
                            total_gen_cur += GenerationCurtailment\
                                    [xh, xt, xco, xn]
                            total_gen_cur_hours[xt] += GenerationCurtailment\
                                    [xh, xt, xco, xn]
                    print()
                print("];")
            print()

            if self.PrintinScreenOptions['OF']:
                if isinstance(obj, EnergyandNetwork):
                    print("\nObjective Function = {}\n".format(\
                        obj.GetObjectiveFunctionENM()))
                elif isinstance(obj, Networkmodel):
                    print("\nObjective Function = {}\n".format(\
                        obj.GetObjectiveFunctionNM()))
            
            print("\nTotal Demand= {}".format(total_demand))
            print("\nTotal Thermal Gen= {}".format(total_thermal_gen))
            print("\nTotal RES Gen= {}".format(total_RES_gen))
            print("\nTotal Hydro Gen= {}".format(total_hydro_gen))
            print("\nTotal Demand Curtailment= {}".format(total_load_curtailment))
            print("\nTotal Generation Curtailment= {}".format(total_gen_cur))
            print("\nTotal Demand hours= {}".format(total_demand_hours))
            print("\nTotal Thermal Gen hours= {}".format(total_thermal_gen_hours))
            print("\nTotal RES Gen hours= {}".format(total_RES_gen_hours))
            print("\nTotal Hydro Gen hours= {}".format(total_hydro_gen_hours))
            print("\nTotal Demand Curtailment hours= {}".format(total_load_curtailment_hours))
            print("\nTotal Generation Curtailment hours= {}".format(total_gen_cur_hours))

    
    def printallEnergyResults(self, obj=None):
        ''' This class method prints on the screen all results for the \
            energy model'''
        if obj is None:
            print('No object with GLPK results has been passed - Aborting \
                printing of results')
            return
        
        PartialStorage = obj.GetPartialStorage()
        TotalStorage = obj.GetTotalStorage()

        if self.PrintinScreenOptions['SoC']:
            for xv in range(obj.NumberTrees):
                print('Vector No:', xv)
                for xtr in range(obj.TreeNodes):
                    print("SoC[%3.0f" % xtr, "][0:1]=[%10.2f"
                          % PartialStorage[xv, xtr], ", %10.2f"
                          % TotalStorage[xv, xtr], "]")
    
    #TODO: This is not well printed with plenty of repetitions, to be discussed 
    # with Alex and Paul
    def PrintallResults(self, obj=None):
        ''' This class method prints on the screen all results for the \
            energy and network model'''
        if obj is None:
            print('No object with GLPK results has been passed - Aborting \
                printing of results')
            return
        if isinstance(obj, EnergyandNetwork) or isinstance(obj, Networkmodel):
            self.printallNetworkResult(obj)
        if isinstance(obj, EnergyandNetwork):
            self.printallEnergyResults(obj)
    
            InputsTree = obj.GetInputsTree()
            OutputsTree = obj.GetOutputsTree()
            print('Water outputs:')
            for xn in range(obj.TreeNodes):
                for xv in range(obj.NumberTrees):
                    print("%8.4f " % OutputsTree[xv, xn], end='')
                print('')
            print('Water inputs:')
            for xn in range(obj.TreeNodes):
                for xv in range(obj.NumberTrees):
                    print("%8.4f " % InputsTree[xv, xn], end='')
                print('')
# -*- coding: utf-8 -*-
"""
Created on Wed Jul 25 11:23:14 2018

Pyene Interfaces provides methods for exchanging information and models with
external tools, such as pypsa and pypower

@author: Dr Eduardo Alejandro Martínez Ceseña
https://www.researchgate.net/profile/Eduardo_Alejandro_Martinez_Cesena
"""
from re import sub
import numpy as np
import logging
import math
import os
import json
import copy
from openpyxl import load_workbook
from typing import Any
from dataclasses import dataclass
from .pyene_Models import network_parameter, model_options_parameter, parameters, tree_parameters, characteristic, information

try:
    import pypsa
except ImportError:
    print('pypsa has not been installed - functionalities unavailable')

from .pyene_Parameters import ElectricityNetwork as ENet


class EInterfaceClass:
    def pyene2pypsa(self, NM, xscen):
        # TODO: Remove xshift and prepare/validate tests
        xshift = 1
        '''Convert pyene files to pypsa format'''
        # Create pypsa network
        try:
            nu = pypsa.Network()
        except ImportError:
            return (0, False)

        nu.set_snapshots(range(NM.settings['NoTime']))
        baseMVA = NM.ENetwork.get_Base()

        # Names
        auxtxtN = 'Bus'
        auxtxtG = 'Gen'
        auxtxtLd = 'Load'

        '''                             NETWORK
        name - network name
        snapshots - list of snapshots or time steps
        snapshot_weightings - weights to control snapshots length
        now - current snapshot
        srid - spatial reference
        '''

        '''                               BUS
        Missing attributes:
            type - placeholder (not yet implemented in pypsa)
            carrier - 'AC' or 'DC'

        Implemented attributes:
            Name auxtxtN+str(xn)
            v_nom - Nominal voltage
            v_mag_pu_set - per unit voltage set point
            v_mag_pu_min - per unit minimum voltage
            v_mag_pu_max - per unit maximum voltage
            auxtxtN = 'Bus'  # Generic name for the nodes
            x - coordinates
            y - coordinates
        '''
        PVBus = np.zeros(NM.ENetwork.get_NoBus(), dtype=float)
        for ob in NM.Gen.Conv:
            if NM.ENetwork.Bus[ob.get_BusPos()].get_Type() == 2:
                PVBus[ob.get_BusPos()] = ob.get_VG()

        for ob in NM.ENetwork.Bus:
            if ob.get_kV() == 0:
                aux1 = 1
            else:
                aux1 = ob.get_kV()
            if ob.get_Type() == 2 or ob.get_Type() == 3:
                aux2 = ob.get_VM()
            else:
                aux2 = PVBus[ob.get_Pos()]

            if ob.get_X() is not None:
                nu.add('Bus', auxtxtN+str(ob.get_Number()),
                       v_nom=aux1, v_mag_pu_set=aux2,
                       v_mag_pu_min=ob.get_Vmin(),
                       v_mag_pu_max=ob.get_Vmax(),
                       x=ob.get_X(),
                       y=ob.get_Y())
            else:
                nu.add('Bus', auxtxtN+str(ob.get_Number()), v_nom=aux1,
                       v_mag_pu_set=aux2, v_mag_pu_min=ob.get_Vmin(),
                       v_mag_pu_max=ob.get_Vmax())

        '''                            GENERATOR
        Missing attributes:
            type - placeholder for generator type
            p_nom_extendable - boolean switch to increase capacity
            p_nom_min - minimum value for extendable capacity
            p_nom_max - maximum value for extendable capacity
            sign - power sign
            carrier - required for global constraints
            capital_cost - cost of extending p_nom by 1 MW
            efficiency - required for global constraints
            committable - boolean (only if p_nom is not extendable)
            start_up_cost - only if commitable is true
            shut_down_cost - only if commitable is true
            min_up_time - only if commitable is true
            min_down_time - only if commitable is true
            initial_status  - only if commitable is true
            ramp_limit_up - only if commitable is true
            ramp_limit_down - only if commitable is true
            ramp_limit_start_up - only if commitable is true
            ramp_limit_shut_down - only if commitable is true

        Implemented attributes:
            name - generator name
            bus - bus name
            control - 'PQ', 'PV' or 'Slack'
            p_min_pu - use default 0 for modelling renewables
            p_max_pu - multpier for intermittent maximum generation
            marginal_cost - linear model
        '''
        # Conventional fuel based generation
        xg = 0
        for ob in NM.Gen.Conv:
            xg += 1
            if NM.ENetwork.Bus[ob.get_BusPos()].get_Type() == 1:
                aux1 = 'PQ'
            elif NM.ENetwork.Bus[ob.get_BusPos()].get_Type() == 2:
                aux1 = 'PV'
            else:
                aux1 = 'Slack'
            nu.add('Generator', auxtxtG+str(xg),
                   bus=auxtxtN+str(ob.get_Bus()),
                   control=aux1,
                   p_nom_max=ob.get_Max()*baseMVA,
                   p_set=ob.get_P(),
                   q_set=ob.get_Q(),
                   marginal_cost=ob.cost['LCost'][0][0]
                   )

        yres = np.zeros(NM.settings['NoTime'], dtype=float)
        for ob in NM.Gen.RES:
            xg += 1
            for xt in range(NM.settings['NoTime']):
                yres[xt] = (NM.scenarios['RES']
                            [NM.resScenario[ob.get_Pos()][xscen]+xt])
            nu.add('Generator', auxtxtG+str(xg+1),
                   bus=auxtxtN+str(ob.get_Bus()),
                   control='PQ',
                   p_nom_max=yres,
                   p_nom=ob.get_Max()*baseMVA,
                   p_set=0,
                   q_set=0,
                   marginal_cost=ob.cost['LCost'][0][0]
                   )

        '''                             CARRIER
        name - Energy carrier
        co2_emissions - tonnes/MWh
        carrier_attribute
        '''

        '''                         GLOBAL CONSTRAINT
        name - constraint
        type - only 'primary_energy' is supported
        carrier_attribute - attributes such as co2_emissions
        sense - either '<=', '==' or '>='
        constant - constant for right hand-side of constraint
        '''

        #                           STORAGE UNIT
        #                              STORE
        '''                              LOAD
        Missing attributes:
            type - placeholder for type
            sign - change to -1 for generator

        Implemented attributes:
            name - auxtxtLd+str(xL)
            bus - Name of bus
        '''
        xL = 0
        ydemP = np.zeros(NM.settings['NoTime'], dtype=float)
        ydemQ = np.zeros(NM.settings['NoTime'], dtype=float)
        for xn in range(NM.ENetwork.get_NoBus()):
            if NM.demandE['PD'][xn] != 0:
                xL += 1
                for xt in range(NM.settings['NoTime']):
                    aux = (NM.scenarios['Demand']
                           [NM.busScenario[xn][xscen]+xt])
                    ydemP[xt] = NM.demandE['PD'][xn]*aux
                    ydemQ[xt] = NM.demandE['QD'][xn]*aux
                nu.add('Load', auxtxtLd+str(xL),
                       bus=auxtxtN+str(xn+1),
                       p_set=ydemP,
                       q_set=ydemQ
                       )

        #                         SHUNT IMPEDANCE

        '''                              LINE
        Missing attributes:
            type - assign string to re-calculate line parameters
            g - shunt conductivity
            s_nom_extendable - boolean switch to allow s_nom to be extended
            s_nom_min - minimum capacity for extendable s_nom
            s_nom_max - maximum capacity for extendable s_nom
            s_max_pu - multiplier (time series) for considering DLR
            capital_cost - cost for extending s_nom by 1
            length - line length
            terrain_factor - terrain factor for increasing capital costs
            num_parallel - number of lines in parallel
            v_ang_min - placeholder for minimum voltage angle drop/increase
            v_and_max - placeholder for maximum voltage angle drop/increase
            b - shunt susceptance

        Implemented attributes:
            Name - auxtxtL+str(xb)
            bus0 - One of the buses connected to the line
            bus1 - The other bus connected to the line
            x - series reactance
            r - series resistance
            snom - MVA limit
        '''
        auxtxtL = 'Line'
        for ob in NM.ENetwork.Branch:
            if ob.get_Tap() == 0:
                auxpu = nu.buses['v_nom']['Bus{}'.format(ob.get_BusF())]**2 / \
                    NM.ENetwork.get_Base()
                nu.add('Line', auxtxtL+str(ob.get_Number()),
                       bus0=auxtxtN+str(ob.get_BusF()),
                       bus1=auxtxtN+str(ob.get_BusT()),
                       x=ob.get_X()*auxpu,
                       r=ob.get_R()*auxpu,
                       s_nom=ob.get_Rate()*NM.ENetwork.get_Base()
                       )

        #                           LINE TYPES
        '''                           TRANSFORMER
        Missing attributes:
            type - assign string to re-calculate line parameters
            g - shunt conductivity
            s_nom_extendable - boolean switch to allow s_nom to be extended
            s_nom_min - minimum capacity for extendable s_nom
            s_nom_max - maximum capacity for extendable s_nom
            s_max_pu - multiplier (time series) for considering DLR
            capital_cost - cost for extending s_nom by 1
            num_parallel - number of lines in parallel
            tap_position - if a type is defined, determine tap position
            phase_shift - voltage phase angle
            v_ang_min - placeholder for minimum voltage angle drop/increase
            v_and_max - placeholder for maximum voltage angle drop/increase

        Implemented attributes:
            Name - auxtxtT+str(xb)
            bus0 - One of the buses connected to the line
            bus1 - The other bus connected to the line
            model - Matpower and pypower use pi admittance models
            tap_ratio - transformer ratio
            tap_side - taps at from bus in Matlab
        '''
        auxtxtT = 'Trs'
        for ob in NM.ENetwork.Branch:
            if ob.get_Tap() != 0:
                nu.add('Transformer', auxtxtT+str(ob.get_Number()),
                       bus0=auxtxtN+str(ob.get_BusF()),
                       bus1=auxtxtN+str(ob.get_BusT()),
                       model='pi',
                       x=ob.get_X(),
                       r=ob.get_R(),
                       b=ob.get_B(),
                       s_nom=ob.get_Rate()*NM.ENetwork.get_Base(),
                       tap_ratio=ob.get_Tap(),
                       tap_side=0
                       )

        #                        TRANSFORMER TYPES
        #                              LINK

        return (nu, True)

    def pypsa2pyene(self):
        '''Convert pyene files to pypsa format'''
        print('To be finalised pypsa2pyene')

    def pyene2pypower(self):
        '''Convert pyene files to pypsa format'''
        print('To be finalised pyene2pypower')

    def pypower2pyene(self):
        '''Convert pyene files to pypsa format'''
        print('To be finalised pypower2pyene')

class PSSErawInterface:
    ''' This class provides functionalities to read raw files from psse'''
    def __init__(self, pathpsseraw = None):
        ''' The init function open the psse raw file if the path \
            is provided'''
        if pathpsseraw != None:
            self.file = open(pathpsseraw, 'r')
        else:
            self.file = None
    
    def detectpsseversion(self):
        ''' This function determines the psse version used to store \
            the raw data '''
        self.psseversion = None
        aux1 = self.filecontent[0].split(',')
        aux1 = int(aux1[2])
        if  aux1 == 31 or aux1 == 33 or aux1 == 34:
            self.psseversion = aux1
        else:
            logging.warning('PSSE version not supported')

    def parsesystemdata(self):
        ''' This function parse the information that corresponds to \
            the whole power system '''
        aux1 = self.filecontent[0].split(',')
        aux2 = aux1[5].split('/')
        print(aux1[1])
        self.BaseUnitPower = float(aux1[1])
        self.BaseFrequency = float(aux2[0])
        self.UnitsTrafosRatings = int(aux1[3]) #  value <= 0 for MVA
                            # value > 0 for current expressed as MVA
        self.UnitsnonTrafosRatings = int(aux1[4]) #  value <= 0 for MVA
                            # value > 0 for current expressed as MVA
    
    def parsenodedata(self):
        ''' This function parse the information that corresponds to \
            the nodes of the power system '''
        aux2 = self.filecontent[3].split(',')
        self.BusNumber = np.array(int(aux2[0]))
        self.BusNomV = np.array(float(aux2[2]))
        aux1 = 4
        while self.filecontent[aux1][0:3] != '0 /':
            aux2 = self.filecontent[aux1].split(',')
            self.BusNumber = np.append(self.BusNumber, int(aux2[0]))
            self.BusNomV = np.append(self.BusNomV, float(aux2[2]))
            aux1 += 1
        self.NumberNodes = len(self.BusNumber)

    
    def parsetrafodata(self):
        ''' This function parse the information that corresponds to \
            the transformers of the power system '''
        aux1 = 14553
        aux3 = 0
        aux4 = 0
        self.ThreeWindingR = np.empty(1)
        self.ThreeWindingX = np.empty(1)
        self.ThreeWindingFrom = np.empty(1)
        self.ThreeWindingTo = np.empty(1)
        self.ConvTrafoR = np.empty(1)
        self.ConvTrafoX = np.empty(1)
        self.ConvTrafoFrom = np.empty(1)
        self.ConvTrafoTo = np.empty(1)        
        while self.filecontent[aux1][0:3] != '0 /':
            aux2 = self.filecontent[aux1].split(',')
            if int(aux2[2]) == 0:
                aux4 += 1
                Bus1 = int(aux2[0])
                Bus2 = int(aux2[1])
                CW = int(aux2[4])
                CZ = int(aux2[5])
                CM = int(aux2[6])
                aux1 += 1
                aux2 = self.filecontent[aux1].split(',')
                R12 = float(aux2[0])
                X12 = float(aux2[1])
                SBase12 = float(aux2[2])
                aux1 += 1
                aux2 = self.filecontent[aux1].split(',')
                NOMV1 = float(aux2[1])
                aux1 += 1
                aux2 = self.filecontent[aux1].split(',')
                NOMV2 = float(aux2[1])
                aux1 += 1
                # Verifying and adapting impedances to base power
                if CZ == 2:
                    if SBase12 != self.BaseUnitPower:
                        R12 = R12 * (self.BaseUnitPower/SBase12)
                        X12 = X12 * (self.BaseUnitPower/SBase12)
                elif CZ == 3:
                    R12 = R12/(self.BaseUnitPower * 1e6)
                    X12 = math.sqrt(X12**2 - R12**2) 
                if aux4 == 1:
                    self.ConvTrafoR[0] = R12
                    self.ConvTrafoX[0] = X12
                    self.ConvTrafoFrom[0] = Bus1
                    self.ConvTrafoTo[0] = Bus2
                else:
                    self.ConvTrafoR = np.append(self.ConvTrafoR, R12)
                    self.ConvTrafoX = np.append(self.ConvTrafoX, X12)
                    self.ConvTrafoFrom = np.append(\
                        self.ConvTrafoFrom, Bus1)
                    self.ConvTrafoTo = np.append(\
                        self.ConvTrafoTo, Bus2)
            else:
                aux3 += 1
                Bus1 = int(aux2[0])
                Bus2 = int(aux2[1])
                Bus3 = int(aux2[2])
                CW = int(aux2[4])
                CZ = int(aux2[5])
                CM = int(aux2[6])
                aux1 += 1
                aux2 = self.filecontent[aux1].split(',')
                R12 = float(aux2[0])
                X12 = float(aux2[1])
                SBase12 = float(aux2[2])
                R23 = float(aux2[3])
                X23 = float(aux2[4])
                SBase23 = float(aux2[5])
                R31 = float(aux2[6])
                X31 = float(aux2[7])
                SBase31 = float(aux2[8])
                aux1 += 1
                aux2 = self.filecontent[aux1].split(',')
                NOMV1 = float(aux2[1])
                aux1 += 1
                aux2 = self.filecontent[aux1].split(',')
                NOMV2 = float(aux2[1])
                aux1 += 1
                aux2 = self.filecontent[aux1].split(',')
                NOMV3 = float(aux2[1])
                aux1 += 1
                # Verifying and adapting impedances to base power
                if CZ == 2:
                    if SBase12 != self.BaseUnitPower:
                        R12 = R12 * (self.BaseUnitPower/SBase12)
                        X12 = X12 * (self.BaseUnitPower/SBase12)
                    if SBase23 != self.BaseUnitPower:
                        R23 = R23 * (self.BaseUnitPower/SBase23)
                        X23 = X23 * (self.BaseUnitPower/SBase23)
                    if SBase31 != self.BaseUnitPower:
                        R31 = R31 * (self.BaseUnitPower/SBase31)
                        X31 = X31 * (self.BaseUnitPower/SBase31)
                elif CZ == 3:
                    R12 = R12/(self.BaseUnitPower * 1e6)
                    X12 = math.sqrt(X12**2 - R12**2)
                    R23 = R23/(self.BaseUnitPower * 1e6)
                    X23 = math.sqrt(X23**2 - R23**2)
                    R31 = R31/(self.BaseUnitPower * 1e6)
                    X31 = math.sqrt(X31**2 - R31**2)

                # Verifying inconsistent reactances and resistances
                minX = min(X12, X23, X31)
                if (X12/minX) > 100:
                    X12 = 100 * minX
                if (X23/minX) > 100:
                    X23 = 100 * minX
                if (X31/minX) > 100:
                    X31 = 100 * minX

                # From delta to star
                R1 = (1/2)*(R12 + R31 - R23)
                R2 = (1/2)*(R12 + R23 - R31)
                R3 = (1/2)*(R31 + R23 - R12)
                X1 = (1/2)*(X12 + X31 - X23)
                X2 = (1/2)*(X12 + X23 - X31)
                X3 = (1/2)*(X31 + X23 - X12)
                if aux3 == 1:
                    self.ThreeWindingR[0] = R1
                    self.ThreeWindingR = np.append(self.ThreeWindingR, R2)
                    self.ThreeWindingR = np.append(self.ThreeWindingR, R3)
                    self.ThreeWindingX[0] = X1
                    self.ThreeWindingX = np.append(self.ThreeWindingX, X2)
                    self.ThreeWindingX = np.append(self.ThreeWindingX, X3)
                    self.ThreeWindingFrom[0] = Bus1
                    self.ThreeWindingFrom = np.append(\
                        self.ThreeWindingFrom, Bus2)
                    self.ThreeWindingFrom = np.append(\
                        self.ThreeWindingFrom, Bus3)
                    self.ThreeWindingTo[0] = 1000000 + aux3
                    self.ThreeWindingTo = np.append(\
                        self.ThreeWindingTo, 1000000 + aux3)
                    self.ThreeWindingTo = np.append(\
                        self.ThreeWindingTo, 1000000 + aux3)
                else:
                    self.ThreeWindingR = np.append(self.ThreeWindingR, R1)
                    self.ThreeWindingR = np.append(self.ThreeWindingR, R2)
                    self.ThreeWindingR = np.append(self.ThreeWindingR, R3)
                    self.ThreeWindingX = np.append(self.ThreeWindingX, X1)
                    self.ThreeWindingX = np.append(self.ThreeWindingX, X2)
                    self.ThreeWindingX = np.append(self.ThreeWindingX, X3)
                    self.ThreeWindingFrom = np.append(\
                        self.ThreeWindingFrom, Bus1)
                    self.ThreeWindingFrom = np.append(\
                        self.ThreeWindingFrom, Bus2)
                    self.ThreeWindingFrom = np.append(\
                        self.ThreeWindingFrom, Bus3)
                    self.ThreeWindingTo = np.append(\
                        self.ThreeWindingTo, 1000000 + aux3)
                    self.ThreeWindingTo = np.append(\
                        self.ThreeWindingTo, 1000000 + aux3)
                    self.ThreeWindingTo = np.append(\
                        self.ThreeWindingTo, 1000000 + aux3)  
        self.NumberThreeWindingTrafos = len(self.ThreeWindingTo)
        self.NumberConvTrafos = len(self.ConvTrafoR)

    def pssedataparser(self):
        ''' This function parse the psse data per device '''
        self.parsesystemdata()

    
    def readpsserawfile(self):
        if self.file == None:
            logging.warning('No PSSE raw file to read')
            return
        self.filecontent = self.file.readlines()
        self.detectpsseversion()
        self.parsesystemdata()
        self.parsenodedata()
        self.parsetrafodata()
        self.file.close()
        with open('pyene/externaldata/datatrafos3.dat', 'w') as f:
            for aux1 in range(self.NumberConvTrafos):
                f.write("%d\t%d\t%.9f\t%.9f\n" %(self.ConvTrafoFrom[aux1], \
                    self.ConvTrafoTo[aux1], self.ConvTrafoR[aux1], \
                    self.ConvTrafoX[aux1]))
            for aux1 in range(self.NumberThreeWindingTrafos):
                f.write("%d\t%d\t%.9f\t%.9f\n" %(self.ThreeWindingFrom[aux1], \
                    self.ThreeWindingTo[aux1], self.ThreeWindingR[aux1], \
                    self.ThreeWindingX[aux1]))
        



        print(self.filecontent[0].split(','))
        print(type(self.filecontent))
        print(type(self.filecontent[0]))

    def openpsserawfile(self, pathpsseraw = None):
        ''' This definition allows opening a psse raw file without passing the \
            path when the object is created'''
        if pathpsseraw != None:
            self.file = open(pathpsseraw, 'r')
        else:
            logging.warning('No path has been pass to the PSSE raw reader\
                therefore none file has been opened')

class any2json:
    
    def matpower2json(self, folder_path=None, name_matpower=None, name_json=None):
        ''' This class method converts matpower files to json files with the 
        format required by pyene
        
        - The path must have the following characteristics:
            * folder_path = path\\to\\folder for windows
            * folder_path = path/to/folder for linux
        - The extension .m and .json must not be included in name_matpower and
        name_json
        '''
        assert folder_path != None, 'No directory path has been pass to load\
            the matpower file'
        assert name_matpower != None, 'No file name has been pass to load\
            the matpower file'
        
        filepath = os.path.join(folder_path, name_matpower+'.m')

        jsonformat = {
            "metadata": {
                "title": name_json,
                "description": 'power system of '+name_json,
                "minimum_version": "0.4"
            },
            "version": 2,
            "baseMVA": None,
            "NoBus": None,
            "NoBranch": None,
            "NoGen": None,
            "Slack": None,
            "bus": {
                "BUS_I": [],
                "BUS_TYPE": [],
                "PD": [],
                "QD": [],
                "GS": [],
                "BS": [],
                "BUS_AREA": [],
                "VM": [],
                "VA": [],
                "BASE_KV": [],
                "ZONE": [],
                "VMAX": [],
                "VMIN": []
            },
            "branch": {
                "F_BUS": [],
                "T_BUS": [],
                "BR_R": [],
                "BR_X": [],
                "BR_B": [],
                "RATE_A": [],
                "RATE_B": [],
                "RATE_C": [],
                "TAP": [],
                "SHIFT": [],
                "BR_STATUS": [],
                "ANGMIN": [],
                "ANGMAX": []
            },
            "gen": {
                "GEN_BUS": [],
                "PG": [],
                "QG": [],
                "QMAX": [],
                "QMIN": [],
                "VG": [],
                "MBASE": [],
                "GEN": [],
                "PMAX": [],
                "PMIN": [],
                "PC1": [],
                "PC2": [],
                "QC1MIN": [],
                "QC1MAX": [],
                "QC2MIN": [],
                "QC2MAX": [],
                "RAMP_AGC": [],
                "RAMP_10": [],
                "RAMP_30": [],
                "RAMP_Q": [],
                "APF": []
            },
            "gencost": {
                "MODEL": [],
                "STARTUP": [],
                "SHUTDOWN": [],
                "NCOST": [],
                "COST": []
            }
        }

        with open(filepath) as fp:
            line = fp.readline()
            flags_bus=False
            flags_gen=False
            flags_branch=False
            flags_gencost=False
            while line:
                if line.split() != [] and line.split()[0] == 'mpc.baseMVA':
                    aux = ""
                    for x in line.split()[2]:
                        if x != ";":
                            aux = aux + x
                    jsonformat['baseMVA'] = float(aux)
                elif line.split() != [] and line.split()[0] == 'mpc.bus':
                    flags_bus = True
                    line = fp.readline()
                    continue
                elif line.split() != [] and line.split()[0] == 'mpc.gen':
                    flags_gen = True
                    line = fp.readline()
                    continue
                elif line.split() != [] and line.split()[0] == 'mpc.branch':
                    flags_branch = True
                    line = fp.readline()
                    continue
                elif line.split() != [] and line.split()[0] == 'mpc.gencost':
                    flags_gencost = True
                    line = fp.readline()
                    continue
                
                if flags_bus and line.split() != [] and line.split()[0] != '];':
                    aux1 = line.split()
                    for val, pos in zip(aux1, jsonformat['bus'].keys()):
                        if pos == 'BUS_TYPE' and int(aux1[1]) == 3:
                            jsonformat['Slack'] = int(aux1[0])
                            jsonformat['bus']['BUS_TYPE'].append(int(aux1[1]))
                        elif pos == 'VMIN':
                            aux2 = ""
                            for x in val:
                                if x != ";":
                                    aux2 = aux2 + x
                            jsonformat['bus']['VMIN'].append(float(aux2))
                        elif pos != 'BUS_I' and pos != 'BUS_TYPE' and \
                            pos != 'BUS_AREA' and pos != 'ZONE':
                            jsonformat['bus'][pos].append(float(val))
                        else:
                            jsonformat['bus'][pos].append(int(val))
                elif flags_bus and line.split() != [] and line.split()[0] == '];':
                    flags_bus = False
                
                if flags_gen and line.split() != [] and line.split()[0] != '];':
                    aux1 = line.split()
                    for val, pos in zip(aux1, jsonformat['gen'].keys()):
                        if pos == 'APF':
                            aux2 = ""
                            for x in val:
                                if x != ";":
                                    aux2 = aux2 + x
                            jsonformat['gen']['APF'].append(float(aux2))
                        elif pos != 'GEN_BUS' and pos != 'GEN':
                            jsonformat['gen'][pos].append(float(val))
                        else:
                            jsonformat['gen'][pos].append(int(val))
                elif flags_gen and line.split() != [] and line.split()[0] == '];':
                    flags_gen = False
                
                if flags_branch and line.split() != [] and line.split()[0] != '];':
                    aux1 = line.split()
                    for val, pos in zip(aux1, jsonformat['branch'].keys()):
                        if pos == 'ANGMAX':
                            aux2 = ""
                            for x in val:
                                if x != ";":
                                    aux2 = aux2 + x
                            jsonformat['branch']['ANGMAX'].append(float(aux2))
                        elif pos != 'F_BUS' and pos != 'T_BUS' and \
                            pos != 'BR_STATUS':
                            jsonformat['branch'][pos].append(float(val))
                        else:
                            jsonformat['branch'][pos].append(int(val))
                elif flags_branch and line.split() != [] and line.split()[0] == '];':
                    flags_branch = False
                
                if flags_gencost and line.split() != [] and line.split()[0] != '];':
                    aux1 = line.split()
                    cnt = 0
                    for pos in jsonformat['gencost'].keys():
                        if pos == 'COST':
                            auxlist = []
                            for x in range(int(aux1[3])):
                                if x < int(aux1[3]) - 1:
                                    auxlist.append(float(aux1[cnt + x]))
                                else:
                                    aux2 = ""
                                    for x1 in aux1[cnt + x]:
                                        if x1 != ";":
                                            aux2 = aux2 + x1
                                    auxlist.append(int(aux2))
                            jsonformat['gencost']['COST'].append(auxlist)
                        elif pos != 'MODEL' and pos != 'NCOST':
                            jsonformat['gencost'][pos].append(float(aux1[cnt]))
                        else:
                            jsonformat['gencost'][pos].append(int(aux1[cnt]))
                        cnt += 1
                elif flags_gencost and line.split() != [] and line.split()[0] == '];':
                    flags_gencost = False                
                line = fp.readline()
        
        jsonformat['NoBus'] = len(jsonformat['bus']['BUS_I'])
        jsonformat['NoBranch'] = len(jsonformat['branch']['F_BUS'])
        jsonformat['NoGen'] = len(jsonformat['gen']['GEN_BUS'])

        filepath = os.path.join(folder_path, name_json+'.json')
        with open(filepath, 'w') as json_file:
            json.dump(jsonformat, json_file, indent=4)

class pyene2any:

    def pyene2matpower(self, folder_path=None, name_matpower=None, ob=None):
        ''' This class method converts pyene data to matpower files with the 
        format required by matpower
        
        - The path must have the following characteristics:
            * folder_path = path\\to\\folder for windows
            * folder_path = path/to/folder for linux
        - The extension .m must not be included in name_matpower
        - The object needs to be a class or subclass of ElectricityNetwork
            of pyene_Parameters.py
        '''

        assert folder_path != None, 'No directory path has been pass to load\
            the matpower file'
        assert name_matpower != None, 'No file name has been pass to load\
            the matpower file'
        assert isinstance(ob, ENet), 'No valid object has been passed'
        
        filepath = os.path.join(folder_path, name_matpower+'.m')

        f= open(filepath,"w+")

        f.write("function mpc = Egypt_reduced\n\n")
        f.write("mpc.version = '2';\n\n")
        f.write("mpc.baseMVA = {0};\n\n".format(\
            ob.get_element(name='baseMVA')))
        f.write("mpc.bus = [\n")
        for x in range(ob.get_no_objects(name='bus')):
            f.write("\t{0}\t{1}\t{2}\t{3}\t{4}\t{5}\t{6}\t{7}\t{8}\t{9}\t{10}\
                \t{11}\t{12};\n".format(\
                ob.get_object_elements(name_object='bus', \
                name_element='number', pos_object=x),
                ob.get_object_elements(name_object='bus', \
                name_element='type', pos_object=x),
                ob.get_object_elements(name_object='bus', \
                name_element='active_power_demand_peak', pos_object=x) * \
                ob.get_element(name='baseMVA'),
                ob.get_object_elements(name_object='bus', \
                name_element='reactive_power_demand_peak', pos_object=x) * \
                ob.get_element(name='baseMVA'),
                0, 0, 0, 0, 0,
                ob.get_object_elements(name_object='bus', \
                name_element='voltage_kv', pos_object=x),
                0,
                ob.get_object_elements(name_object='bus', \
                name_element='maximum_voltage_magnitude', pos_object=x),
                ob.get_object_elements(name_object='bus', \
                name_element='minimum_voltage_magnitude', pos_object=x)))
        f.write("];\n\n")

        f.write("mpc.gen = [\n")
        for x in range(ob.get_no_objects(name='conv')):
            f.write("\t{0}\t{1}\t{2}\t{3}\t{4}\t{5}\t{6}\t{7}\t{8}\t{9}\t{10}\
                \t{11}\t{12}\t{13}\t{14}\t{15}\t{16}\t{17}\t{18}\t{19}\t{20}\
                ;\n".format(\
                ob.get_object_elements(name_object='conv', \
                name_element='bus_number', pos_object=x),
                0, 0,
                ob.get_object_elements(name_object='conv', \
                name_element='maximum_reactive_power_generation', pos_object=x),
                ob.get_object_elements(name_object='conv', \
                name_element='minimum_reactive_power_generation', pos_object=x),
                0, 0,
                ob.get_object_elements(name_object='conv', \
                name_element='status', pos_object=x),
                ob.get_object_elements(name_object='conv', \
                name_element='maximum_active_power_generation', pos_object=x) \
                * ob.get_element(name='baseMVA'),
                ob.get_object_elements(name_object='conv', \
                name_element='minimum_active_power_generation', pos_object=x) \
                * ob.get_element(name='baseMVA'),
                0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0))
        f.write("];\n\n")

        f.write("mpc.branch = [\n")
        for x in range(ob.get_no_objects(name='transmissionline')):
            aux = ob.get_object_elements(name_object='transmissionline', \
                name_element='bus_number', pos_object=x)
            f.write("\t{0}\t{1}\t{2}\t{3}\t{4}\t{5}\t{6}\t{7}\t{8}\t{9}\t{10}\
                \t{11}\t{12};\n".format(\
                aux[0], aux[1],
                ob.get_object_elements(name_object='transmissionline', \
                name_element='resistance', pos_object=x),
                ob.get_object_elements(name_object='transmissionline', \
                name_element='reactance', pos_object=x),
                ob.get_object_elements(name_object='transmissionline', \
                name_element='shunt_susceptance', pos_object=x),
                ob.get_object_elements(name_object='transmissionline', \
                name_element='long_term_thermal_limit', pos_object=x) \
                * ob.get_element(name='baseMVA'),
                0, 0, 0, 0,
                ob.get_object_elements(name_object='transmissionline', \
                name_element='status', pos_object=x),
                0, 0))
        for x in range(ob.get_no_objects(name='transformers')):
            aux = ob.get_object_elements(name_object='transformers', \
                name_element='bus_number', pos_object=x)
            f.write("\t{0}\t{1}\t{2}\t{3}\t{4}\t{5}\t{6}\t{7}\t{8}\t{9}\t{10}\
                \t{11}\t{12};\n".format(\
                aux[0], aux[1],
                ob.get_object_elements(name_object='transformers', \
                name_element='resistance', pos_object=x),
                ob.get_object_elements(name_object='transformers', \
                name_element='reactance', pos_object=x),
                ob.get_object_elements(name_object='transformers', \
                name_element='shunt_susceptance', pos_object=x),
                ob.get_object_elements(name_object='transformers', \
                name_element='long_term_thermal_limit', pos_object=x) \
                * ob.get_element(name='baseMVA'),
                0, 0,
                ob.get_object_elements(name_object='transformers', \
                name_element='tap', pos_object=x),
                0,
                ob.get_object_elements(name_object='transformers', \
                name_element='status', pos_object=x),
                -360, 360))
        f.write("];\n\n")

        f.write("mpc.gencost = [\n")
        for x in range(ob.get_no_objects(name='conv')):
            f.write("\t{0}\t{1}\t{2}\t{3}".format(\
                2, 0, 0,
                len(ob.get_object_elements(name_object='conv', \
                name_element='cost_function_parameters', pos_object=x)),
                ))
            aux = ob.get_object_elements(name_object='conv', \
                name_element='cost_function_parameters', pos_object=x)
            for y in aux:
                f.write("\t{0}".format(y))
            f.write(";\n")

        f.write("];\n\n")

        f.close()

@dataclass
class tree_parameters_read:
    name                :   str     = None      # Name parameter
    level               :   int     = None      # Level in the tree
    pos                 :   int     = None      # Position in level
    value               :   Any   = None      # Value of parameter

@dataclass
class profile_parameter:
    name                :   str     = None      # Name of the parameter
    position_tree       :   dict    = None      # Position in the energy tree - representative days
                                                # in case of parameters changing in time
    hour                :   int     = None      # Hour of the parameter in case of parameters 
                                                # changing in time
    ID                  :   str     = None      # ID of element
    type                :   str     = None      # Type of element, e.g. bus, branch                                                # related
    subtype             :   str     = None      # subtype of element. e.g. conventional, transmission line
    group               :   Any     = None
    zone                :   Any     = None
    value               :   Any     = None      # Value of specific parameter

class excel2pyene:
    ''' This class reads excel files and store the data in an object '''
    
    integer_characteristics_nodes = [
        "number",
        "typePF"
    ]

    double_characteristics_nodes =[    
    ]

    string_characteristics_nodes =[
        "ID",
        "name_bus",
        "zone",
        "group",
        "subtype"
    ]

    bool_characteristics_nodes = []

    integer_parameters_nodes =[
    ]

    double_parameters_nodes =[
        "Pd",
        "Qd",
        "Gs",
        "Bs",
        "baseV",
        "Vmpr",
        "Vapr",
        "Vmax",
        "Vmin"
    ]

    string_parameters_nodes =[
    ]

    bool_parameters_nodes =[]

    #      NAMES ELEMENTS BRANCHES

    integer_characteristics_branches =[
        "from",
        "to"
    ]

    double_characteristics_branches =[
    ]

    string_characteristics_branches =[
        "ID",
        "subtype",
        "group"
    ]

    bool_characteristics_branches =[
    ]

    integer_parameters_branches =[
    ]

    double_parameters_branches =[
        "resistance",
        "reactance",
        "LCsusceptance",
        "maxPflow",
        "CTEP",
        "Vmpr",
        "Vapr",
        "Vmax",
        "Vmin"
    ]

    string_parameters_branches =[
    ]

    bool_parameters_branches =[
        "status",
        "vTEP"
    ]

    #      NAMES ELEMENTS GENERATORS

    integer_characteristics_generators =[
        "number"
    ]

    double_characteristics_generators =[
    ]

    string_characteristics_generators =[
        "ID",
        "subtype",
        "group"
    ]

    bool_characteristics_generators =[
    ]

    integer_parameters_generators =[
    ]

    double_parameters_generators =[
        "Pmax",
        "Pmin",
        "Pfix",
        "Qmax",
        "Qmin",
        "Qfix",
        "cUC",
        "cGEP",
        "fCPg",
        "vCPg",
        "emissions",
        "startup",
        "shutdown",
        "cost function"
    ]

    string_parameters_generators =[
    ]

    bool_parameters_generators =[
        "status",
        "vUC",
        "vGEP"
    ]

    integer_characteristics = copy.copy(integer_characteristics_nodes)
    integer_characteristics.extend(integer_characteristics_branches)
    integer_characteristics.extend(integer_characteristics_generators)
    integer_characteristics = list(dict.fromkeys(integer_characteristics))

    double_characteristics = copy.copy(double_characteristics_nodes)
    double_characteristics.extend(double_characteristics_branches)
    double_characteristics.extend(double_characteristics_generators)
    double_characteristics = list(dict.fromkeys(double_characteristics))

    string_characteristics = copy.copy(string_characteristics_nodes)
    string_characteristics.extend(string_characteristics_branches)
    string_characteristics.extend(string_characteristics_generators)
    string_characteristics = list(dict.fromkeys(string_characteristics))

    bool_characteristics = copy.copy(bool_characteristics_nodes)
    bool_characteristics.extend(bool_characteristics_branches)
    bool_characteristics.extend(bool_characteristics_generators)
    bool_characteristics = list(dict.fromkeys(bool_characteristics))

    all_characteristics = copy.copy(integer_characteristics)
    all_characteristics.extend(double_characteristics)
    all_characteristics.extend(string_characteristics)
    all_characteristics.extend(bool_characteristics)
    all_characteristics = list(dict.fromkeys(all_characteristics))

    integer_parameters = copy.copy(integer_parameters_nodes)
    integer_parameters.extend(integer_parameters_branches)
    integer_parameters.extend(integer_parameters_generators)

    double_parameters = copy.copy(double_parameters_nodes)
    double_parameters.extend(double_parameters_branches)
    double_parameters.extend(double_parameters_generators)

    string_parameters = copy.copy(string_parameters_nodes)
    string_parameters.extend(string_parameters_branches)
    string_parameters.extend(string_parameters_generators)

    bool_parameters = copy.copy(bool_parameters_nodes)
    bool_parameters.extend(bool_parameters_branches)
    bool_parameters.extend(bool_parameters_generators)

    # Predefined characteristics (names, parameters) that are used in the energy engine
    problems_names_nodes = {
        "substation expansion planning" : "SEP",
        "sep" : "SEP"
    }
    problems_names_branches = {
        "transmission expansion planning" : "TEP",
        "tep" : "TEP"
    }
    problems_names_generators = {
        "unit commitment" : "UC",
        "uc" : "UC",
        "generation expansion planning" : "GEP",
        "gep" : "GEP"
    }
    problems_names_system = {
        "dc economic dispatch" : "DC ED",
        "economic dispatch" : "DC ED",
        "dc ed" : "DC ED",
        "dc optimal power flow" : "DC OPF",
        "optimal power flow" : "DC OPF",
        "dc opf" : "DC OPF",
        "ac power flow" : "AC PF",
        "power flow" : "AC PF",
        "ac pf" : "AC PF",
        "network reduction" : "NetR",
        "netr" : "NetR",
        "balance tree" : "BT",
        "bt" : "BT"
    }
    accepted_engines = {
        "pyene" : "pyene",
        "fdif"  : "fdif"
    }
    
    # Parameters nodes
    parameters_nodes_integer = {
        "number" : "number",
        "type node power flow" : "typePF",
        "typepf" : "typePF",
    }
    parameters_nodes_double = {
        "active power demand" : "Pd",
        "pd" : "Pd",
        "reactive power demand" : "Qd",
        "qd" : "Qd",
        "shunt conductance" : "Gs",
        "gs" : "Gs",
        "shunt susceptance" : "Bs",
        "bs" : "Bs",
        "basev" : "baseV",
        "predefined voltage magnitude" : "Vmpr",
        "vmpr" : "Vmpr",
        "predefined voltage angle" : "Vapr",
        "vapr" : "Vapr",
        "max voltage magnitude" : "Vmax",
        "vmax": "Vmax",
        "min voltage magnitude" : "Vmin",
        "vmin": "Vmin"
    }
    parameters_nodes_string = {
        "id" : "ID",
        "name" : "name_bus",
        "zone" : "zone",
        "group" : "group"
    }
    parameters_nodes_bool = { }
    parameters_nodes = {}
    parameters_nodes.update(parameters_nodes_integer)
    parameters_nodes.update(parameters_nodes_double)
    parameters_nodes.update(parameters_nodes_string)
    parameters_nodes.update(parameters_nodes_bool)

    parameters_nodes_list_integer_type = {}
    parameters_nodes_list_double_type = {}
    parameters_nodes_list_string_type = {
        "subtype" : "subtype"
    }
    parameters_nodes_list_bool_type = {}
    parameters_nodes_list_type = {}
    parameters_nodes_list_type.update(parameters_nodes_list_integer_type)
    parameters_nodes_list_type.update(parameters_nodes_list_double_type)
    parameters_nodes_list_type.update(parameters_nodes_list_string_type)
    parameters_nodes_list_type.update(parameters_nodes_list_bool_type)

    parameters_names_nodes = parameters_nodes.copy()
    parameters_names_nodes.update(parameters_nodes_list_type)

    # Parameters branches
    parameters_branches_integer = {
        "from" : "from",
        "to" : "to",
    }
    parameters_branches_double = {
        "length" : "length",
        "resistance" : "resistance",
        "r" : "resistance",
        "reactance" : "reactance",
        "x" : "reactance",
        "line charging susceptance" : "LCsusceptance",
        "shunt susceptance" : "LCsusceptance",
        "b" : "LCsusceptance",
        "max active power flow" : "maxPflow",
        "max pflow" : "maxPflow",
        "cost new branch" : "CTEP",
    }
    parameters_branches_string = {
        "id" : "ID",
        "group" : "group"
    }
    parameters_branches_bool = {
        "status" : "status",
        "tep variable" : "vTEP"
    }
    parameters_branches={}
    parameters_branches.update(parameters_branches_integer)
    parameters_branches.update(parameters_branches_double)
    parameters_branches.update(parameters_branches_string)
    parameters_branches.update(parameters_branches_bool)

    parameters_branches_list_integer_type = {}
    parameters_branches_list_double_type = {}
    parameters_branches_list_bool_type = {}
    parameters_branches_list_string_type = {
        "subtype" : "subtype"
    }
    parameters_branches_list_type = {}
    parameters_branches_list_type.update(parameters_branches_list_integer_type)
    parameters_branches_list_type.update(parameters_branches_list_double_type)
    parameters_branches_list_type.update(parameters_branches_list_bool_type)
    parameters_branches_list_type.update(parameters_branches_list_string_type)

    parameters_names_branches = parameters_branches.copy()
    parameters_names_branches.update(parameters_branches_list_type)
    accepted_types_branches = {
        "ac transmission line" : "TL",
        "transmission line" : "TL",
        "tl" : "TL",
        "ac interconnector" : "inter",
        "interconnector" : "inter",
        "inter" : "inter",
        "transformer" : "trafo",
        "trafo" : "trafo",
        "user" : "user"
    }

    # Parameters generators
    parameters_generators_integer = {
        "number" : "number",
        "pieces" : "pieces",
        "model"  : "model"
    }
    parameters_generators_double = {
        "active power max limit" : "Pmax",
        "pmax" : "Pmax",
        "active power min limit" : "Pmin",
        "pmin" : "Pmin",
        "fixed active power" : "Pfix",
        "pfix" : "Pfix",
        "reactive power max limit" : "Qmax",
        "qmax" : "Qmax",
        "reactive power min limit" : "Qmin",
        "qmin" : "Qmin",
        "fixed reactive power" : "Qfix",
        "qfix" : "Qfix",
        "cost unit commitment" : "cUC",
        "cuc" : "cUC",
        "cost new generator" : "cGEP",
        "cgep" : "cGEP",
        "fixed operation cost" : "fCPg",
        "fcpg" : "fCPg",
        "variable operation cost" : "vCPg",
        "vcpg" : "vCPg",
        "emissions" : "emissions",
        "startup" : "startup",
        "shutdown" : "shutdown"
    }
    parameters_generators_string = {
        "id" : "ID",
        "group" : "group"
    }
    parameters_generators_bool = {
        "status" : "status",
        "unit commitment" : "vUC",
        "uc" : "vUC",
        "gep variable" : "vGEP",
        "vgep" : "vGEP"
    }
    parameters_generators = {}
    parameters_generators.update(parameters_generators_integer)
    parameters_generators.update(parameters_generators_double)
    parameters_generators.update(parameters_generators_string)
    parameters_generators.update(parameters_generators_bool)

    parameters_generators_list_integer_type = {}
    parameters_generators_list_double_type = {
        "cost function" : "cost function"
    }
    parameters_generators_list_string_type = {
        "subtype" : "subtype"
    }
    parameters_generators_list_bool_type = {}
    parameters_generators_list_type = {}
    parameters_generators_list_type.update(parameters_generators_list_integer_type)
    parameters_generators_list_type.update(parameters_generators_list_double_type)
    parameters_generators_list_type.update(parameters_generators_list_string_type)
    parameters_generators_list_type.update(parameters_generators_list_bool_type)

    parameters_names_generators = parameters_generators.copy()
    parameters_names_generators.update(parameters_generators_list_type)
    accepted_types_generators = {
        "thermal" : "thermal",
        "hydro" : "hydro",
        "wind" : "wind",
        "solar" : "solar",
        "user" : "user",
        "diesel" : "diesel"
    }

    # Parameters system
    accepted_characteristics = {
        "lossess"                   : "loss",
        "solver"                    : "solver",
        "base power"                : "Sbase",
        "multiperiod"               : "multiperiod",
        "output file name"          : "output file name",
        "moea"                      : "MOEA",
        "representative periods"    : "representative periods"
    }
    accepted_solvers = {
        "glpk"      :   "GLPK",
        "clp"       :   "CLP",
        "clp-i"     :   "CLP-I",
        "clp-ir"    :   "CLP-IR"
    }

    # Profiles parameters
    unique_string_parameters_profile = {
        "type" : "type",
        "name": "name",
        "id": "ID",
        "group" : "group",
        "zone" : "zone" 
    }
    list_string_parameters_profile = {
        "subtype": "subtype",
        "position tree": "pt",
        "representative day": "pt",
        "pt": "pt"
    }
    parameters_profile = {}
    parameters_profile.update(unique_string_parameters_profile)
    parameters_profile.update(list_string_parameters_profile)
    
    # Parameters connections
    unique_string_parameters_connections = {
        "type" : "type",
        "id": "ID",
        "group" : "group",
        "zone" : "zone"
    }
    list_string_parameters_connections = {
        "problems" : "problems",
        "variables" : "variables",
        "subtype": "subtype",
        "position tree": "pt",
        "representative day": "pt",
        "pt": "pt"
    }

    # Parameters outputs
    unique_integer_parameters_outputs = {
        "hour" : "hour"
    }

    unique_string_parameters_outputs = {
        "type" : "type",
        "id": "ID",
        "group" : "group",
        "zone" : "zone",
        "name" : "name",
        "problem" : "problem",
        "reference" : "reference",
        "data device" : "data device",
        "type information" : "type information",
        "function" : "function",
    }
    list_string_parameters_outputs = {
        "position tree": "pt",
        "representative day": "pt",
        "pt": "pt",
        "subtype": "subtype",
    }

    # Concatenate parameters types
    parameters_integer = {}
    parameters_integer.update(parameters_nodes_integer)
    parameters_integer.update(parameters_branches_integer)
    parameters_integer.update(parameters_generators_integer)
    parameters_integer.update(unique_integer_parameters_outputs)

    parameters_double = {}
    parameters_double.update(parameters_nodes_double)
    parameters_double.update(parameters_branches_double)
    parameters_double.update(parameters_generators_double)

    parameters_string = {}
    parameters_string.update(parameters_nodes_string)
    parameters_string.update(parameters_branches_string)
    parameters_string.update(parameters_generators_string)
    parameters_string.update(problems_names_nodes)
    parameters_string.update(problems_names_branches)
    parameters_string.update(problems_names_generators)
    parameters_string.update(problems_names_system)
    parameters_string.update(unique_string_parameters_profile)
    parameters_string.update(unique_string_parameters_connections)
    parameters_string.update(unique_string_parameters_outputs)

    parameters_bool = {}
    parameters_bool.update(parameters_nodes_bool)
    parameters_bool.update(parameters_branches_bool)
    parameters_bool.update(parameters_generators_bool)

    parameters_list_integer = {}
    parameters_list_integer.update(parameters_nodes_list_integer_type)
    parameters_list_integer.update(parameters_generators_list_integer_type)
    parameters_list_integer.update(parameters_branches_list_integer_type)

    parameters_list_double = {}
    parameters_list_double.update(parameters_nodes_list_double_type)
    parameters_list_double.update(parameters_generators_list_double_type)
    parameters_list_double.update(parameters_branches_list_double_type)
    
    parameters_list_string = {}
    parameters_list_string.update(parameters_nodes_list_string_type)
    parameters_list_string.update(parameters_generators_list_string_type)
    parameters_list_string.update(parameters_branches_list_string_type)
    parameters_list_string.update(list_string_parameters_profile)
    parameters_list_string.update(list_string_parameters_connections)
    parameters_list_string.update(list_string_parameters_outputs)

    parameters_list_bool = {}
    parameters_list_bool.update(parameters_nodes_list_bool_type)
    parameters_list_bool.update(parameters_generators_list_bool_type)
    parameters_list_bool.update(parameters_branches_list_bool_type)

    all_unique_parameters = {}
    all_unique_parameters.update(parameters_integer)
    all_unique_parameters.update(parameters_double)
    all_unique_parameters.update(parameters_string)
    all_unique_parameters.update(parameters_bool)

    all_list_parameters = {}
    all_list_parameters.update(parameters_list_integer)
    all_list_parameters.update(parameters_list_double)
    all_list_parameters.update(parameters_list_string)
    all_list_parameters.update(parameters_list_bool)

    all_parameters = {}
    all_parameters.update(all_unique_parameters)
    all_parameters.update(all_list_parameters)

    # Variables names for power system problems
    variables_balance_tree = [
        "input",
        "output",
        "flow",
        "surplus",
        "deficit"
    ]
    variables_OPF = [
        "active power generation",
        "active power generation cost",
        "active power flow",
        "voltage angle",
        "generation curtailment",
        "load curtailment"
    ]


    def _read_bool_excel(self, parameters_list=[], position=-1, sheet=None, name_compare=""):
        ''' This function reads boolean information on a worksheet

            Parameters
            ----------
            Mandatory: \\
            parameters_list : List of parameters in the excel sheet \\
            position : Row in the excel sheet where the boolean elements are \\
            sheet : excel sheet to be read \\
            name_compare : name of the column to be read
        '''
        elements=[]
        for num, name in enumerate(parameters_list):
            counter = position
            value = sheet.cell(row=counter, column=num+1).value
            while name == name_compare and value is not None:
                new_val = False
                if type(value) == str and value.lower() == "=false" :
                    new_val = False
                elif type(value) == str and value.lower() == "false":
                    new_val = False
                elif type(value) == int and value == 0:
                    new_val = False
                elif type(value) == str and value.lower() == "=true":
                    new_val = True
                elif type(value) == str and value.lower() == "true":
                    new_val = True
                elif type(value) == int and value == 1:
                    new_val = True
                elif type(value) == bool:
                    new_val = value
                elements.append(new_val)
                counter = counter + 1
                value = sheet.cell(row=counter, column=num+1).value
            if elements:
                return elements
        if not elements:
            raise ValueError('Problem retrieving the boolean information')

    def _read_elements_excel(self, parameters_list=[], position=-1, sheet=None, name_compare="", references=[]):
        ''' This function reads information (columns) from an excel sheet

            Parameters
            ----------
            Mandatory: \\
            parameters_list : List of parameters in the excel sheet \\
            position : Row in the excel sheet where the boolean elements are \\
            sheet : excel sheet to be read \\
            name_compare : name of the column to be read\\
            references : List of of valid inputs to compare with the values in the excel sheet 
        '''
        elements = []
        invalid = []
        for num, name in enumerate(parameters_list):
            counter = position
            value = sheet.cell(row=counter, column=num+1).value
            while value is not None and name == name_compare:
                if references and references.get(value.lower().replace(u'\xa0',' '), None) is None:
                    invalid.append(value)
                    elements.append(None)
                elif references and references.get(value.lower().replace(u'\xa0',' '), None) is not None:
                    elements.append(references.get(value.lower().replace(u'\xa0',' '), None))
                elif not references:
                    elements.append(value)
                counter = counter + 1
                value = sheet.cell(row=counter, column=num+1).value
            if invalid:
                # Printing invalid options
                for k in invalid:
                    print('{} is not a valid option'.format(k))
            if elements:
                return elements

    def _revise_model_problem_exist(self, model=None, problem=None, active=None, engine=None):
        ''' This function determines if the problem already exist and update the value in case a new value (active) is given

            Parameters
            ----------
            Mandatory: \\
            model : python object containing the mathematical model and all parameters\\
            problem : name of problem. e.g. TEP, GEP\\
            active : Boolean flag indicating if the problem is active\\
            engine : Indicates if the problem is solved in the integrated framework or in pyene 
        '''
        exist = False
        for element in model.model_options:
            if element.name == problem:
                print("problem {} already exist in model options with value {} and engine {}. The value will be change to {} for engine {}". format(problem, element.value, element.engine, active, engine))
                element.value = active
                element.engine = engine
                exist = True
                break
        return exist
    
    def _revise_model_characteristic_exist(self, model=None, characteristic=None, option=None):
        ''' This function determines if the characteristic already exist and update the value in case a new value (option) is given

            Parameters
            ----------
            Mandatory: \\
            model : Python object containing the mathematical model and all parameters\\
            characteristic : Name of characteristic. e.g. losses, solver\\
            option : option to be given to the chatacteristic
        '''
        exist = False
        for element in model.model_options:
            if element.name == characteristic:
                print("characteristic {} already exist in model options with value {}. The value will be change to {}". format(characteristic, element.value, option))
                element.value = option
                exist = True
                break
        return exist
    
    def _read_model_options(self, model=None, sheet=None):
        ''' This function reads the options used in both the integrated framework and pyene for energy optimisation

            Parameters
            ----------
            Mandatory:
            model : python object containing the mathematical model and all parameters\\
            sheet : excel sheet to be read
        '''
        ### Extracting the options from the excel file ###

        parameters_list = []
        counter = 1
        value = sheet.cell(row=2, column=counter).value
        if isinstance(value, str):
            value = value.replace(u'\xa0',' ')
        while value is not None:
            parameters_list.append(value.lower())
            counter = counter + 1
            value = sheet.cell(row=2, column=counter).value
            if isinstance(value, str):
                value = value.replace(u'\xa0',' ')
        
        names = self.problems_names_system.copy()
        names.update(self.problems_names_nodes)
        names.update(self.problems_names_branches)
        names.update(self.problems_names_generators)

        active = self._read_bool_excel(parameters_list, 3, sheet, "active")
        problem = self._read_elements_excel(parameters_list, 3, sheet, "problem", names)
        engine = self._read_elements_excel(parameters_list, 3, sheet, "engine", self.accepted_engines)
        characteristics = self._read_elements_excel(parameters_list, 3, sheet, "characteristics", self.accepted_characteristics)
        options = self._read_elements_excel(parameters_list, 3, sheet, "options")
        files = self._read_elements_excel(parameters_list, 3, sheet, "paths")
        if files is not None:
            self.files2open.append(files)

        if len(problem) == len(active) and len(problem) == len(engine):
            for pro, act, eng in zip(problem, active, engine):
                if not self._revise_model_problem_exist(model, pro, act, eng):
                    model.model_options.append(model_options_parameter(name=pro, value=act, engine=eng))
        else:
            raise ValueError('indicated elements with different sizes.\nproblems = {}\nactive = {}\nengine = {}'.format(len(problem), len(active), len(engine)))
        
        if len(characteristics) == len(options):
            for cha, opt in zip(characteristics, options):
                if not self._revise_model_characteristic_exist(model, cha, opt):
                    model.model_options.append(model_options_parameter(name=cha, value=opt))
        else:
            raise ValueError('indicated elements with different sizes.\ncharacteristics = {}\noptions = {}'.format(len(characteristics), len(options)))

    def _extract_parameters_names(self, parameters_names, sheet):
        ''' This function extracts the information of the titles of the parameters for the nodes

            Parameters
            ----------
            Mandatory:
            sheet:  openpyxl object containing the excel sheet
            parameters_names : List of accepted names depending on the type of element
        '''

        # Extracting the titles of the node information
        parameters_list = []
        unvalid_parameters = []
        counter = 1
        col_id = None
        value = sheet.cell(row=2, column=counter).value
        if isinstance(value, str):
            value = value.replace(u'\xa0',' ')
        while value is not None:
            if value.lower() == "id":
                col_id = counter
            parameters_list.append(parameters_names.get(value.lower()))
            if parameters_names.get(value.lower()) is None:
                unvalid_parameters.append(value)
            counter = counter + 1
            value = sheet.cell(row=2, column=counter).value
            if isinstance(value, str):
                value = value.replace(u'\xa0',' ')

        # Printing the parameters that were indicated in the excel file and that
        # do not belong to the list of accepted parameters
        for k in unvalid_parameters:
            print('{} is not included in the list of parameters that the software can solve'.format(k))
        
        if col_id is None:
            raise ValueError('No column with ID has been identified check that all your worksheets for generators, buses and branches have an ID')

        return parameters_list, col_id

    def _revise_parameters(self, list2check=None, ID=None, name=None, value=None):
        ''' This function check if a parameter exists and update the value. An exception is made with the parameter subtype to create as many subtypes as the user wants

            Parameters
            ----------
            Mandatory:\\
            list2check : list containing the parameters to be analysed
            ID : ID of element to be search in the list of parameters
            name : Name of the parameter
            value : Value of the parameter
        '''
        if name == "subtype":
            return False
        
        for element in list2check:
            if element.ID == ID and element.name == name:
                print("WARNING! parameter with name {}, type {} and ID {} already exist with value {}. The value is being updated to {}. It is very likely that the same parameter is being loaded from two or more files".format(name, element.type, ID, element.value, value))
                element.value = value
                return True
        return False

    def _read_info_network(self, model=None, sheet=None, parameters_names=None, typ=None):
        ''' This function reads the information of buses generators and branches used in the integrated framework for energy optimisation

            Parameters
            ----------
            Mandatory:\\
            model : python object containing the mathematical model and all parameters\\
            sheet : excel sheet to be read\\
            parameters_names : parameters names that are accepted for each type of element\\
            typ : type of element to be stored. bus, branch or generator
        '''

        ### Extracting the options from the excel file ###
        parameters_list, col_id = self._extract_parameters_names(\
            parameters_names, sheet)
        # Extracting the ID of all elements
        counter = 3
        value_id = sheet.cell(row=counter, column=col_id).value
        all_elements = []
        model.data.data[typ] = []
        while value_id:
            element_all_info = []
            characteristics = information()
            characteristics.characteristics.append(characteristic(name="ID", value=value_id, data_type="string"))
            characteristics.characteristics.append(characteristic(name="type", value=typ, data_type="string"))
            for col, name in enumerate(parameters_list):
                value = sheet.cell(row=counter, column=col + 1).value
                # if col + 1 != col_id and not self._revise_parameters(model.network_parameters, value_id, name, value) and name is not None:
                if col + 1 != col_id and name is not None and value is not None:
                    if name in self.integer_characteristics:
                        characteristics.characteristics.append(characteristic(name=name, value=value, data_type="integer"))
                    elif name in self.double_characteristics:
                        characteristics.characteristics.append(characteristic(name=name, value=value, data_type="double"))
                    elif name in self.bool_characteristics:
                        characteristics.characteristics.append(characteristic(name=name, value=value, data_type="bool"))
                    elif name in self.string_characteristics:
                        if name != "subtype":
                            characteristics.characteristics.append(characteristic(name=name, value=str(value), data_type="string"))
                        elif name == "subtype" and characteristics.exist("subtype"):
                            new_value = characteristics.get_characteristic("subtype")
                            new_value.append(str(value))
                        elif name == "subtype" and not characteristics.exist("subtype"):
                            characteristics.characteristics.append(characteristic(name=name, value=[str(value)], data_type="v_string"))
                    elif name in self.integer_parameters:
                        element_all_info.append(information(data_type="integer", value=value, characteristics=[characteristic(name="name", value=name, data_type="string")]))
                    elif name in self.double_parameters:
                        element_all_info.append(information(data_type="double", value=value, characteristics=[characteristic(name="name", value=name, data_type="string")]))
                    elif name in self.bool_parameters:
                        element_all_info.append(information(data_type="bool", value=value, characteristics=[characteristic(name="name", value=name, data_type="string")]))
                    elif name in self.string_parameters:
                        element_all_info.append(information(data_type="string", value=str(value), characteristics=[characteristic(name="name", value=name, data_type="string")]))
                        
                    #model.network_parameters.append(network_parameter(ID=value_id, type=typ, name=name, value=value))
            for element in element_all_info:
                element.characteristics.extend(characteristics.characteristics)
            
            all_elements.extend(element_all_info)
            
            counter = counter + 1
            value_id = sheet.cell(row=counter, column=col_id).value
        model.data.data[typ] = np.array(all_elements)

    def _revise_tree_info(self, name=None, pos=None, value=None, level=None):
        ''' This function check if a parameter exists and update the value.

            Parameters
            ----------
            Mandatory:\\
            pos: Position in the level\\
            name : Name of the parameter\\
            value : Value of the parameter\\
            level: level in the balance tree
        '''
        for parameter in self.tree_info:
            if parameter.name == name and parameter.level == level and parameter.pos == pos:
                print("WARNING! Parameter {} in level {} and with position {} in that level has a value of {} and the value will be changed to {}".format(name, level, pos, parameter.value, value))
                parameter.value = value
                return False
        return True

    def _read_tree_parameter(self, sheet=None, rows=None, name=None):
        ''' This function reads a specific parameter used in the balance tree model

            Parameters
            ----------
            Mandatory:
            model : python object containing the mathematical model and all parameters\\
            sheet : excel sheet to be read\\
            name: Name of the parameter to be read
        '''
        value = 1
        rows = rows + 1
        columns = 1
        value = sheet.cell(row=rows, column=1).value
        pos = 1
        while value is not None:
            level = value
            while value is not None:
                columns = columns + 1
                value = sheet.cell(row=rows, column=columns).value
                if value is not None and isinstance(value, str) and self._revise_tree_info(name, pos, value.lower(), level):
                    self.tree_info.append(tree_parameters_read(name=name, level=level, pos=pos, value=value.lower()))
                    pos = pos + 1
                elif value is not None and not isinstance(value, str) and self._revise_tree_info(name, pos, value, level):
                    self.tree_info.append(tree_parameters_read(name=name, level=level, pos=pos, value=value))
                    pos = pos + 1
            rows = rows + 1
            columns = 1
            value = sheet.cell(row=rows, column=1).value
            pos = 1
        return rows

    def _read_info_tree(self, sheet=None):
        ''' This function reads all parameters used in the balance tree model

            Parameters
            ----------
            Mandatory:\\
            model : python object containing the mathematical model and all parameters\\
            sheet : excel sheet to be read
        '''
        rows = 2
        tries = 0
        while tries < 20:
            value = sheet.cell(row=rows, column=1).value
            if isinstance(value, str) and value.lower() == "names":
                rows = self._read_tree_parameter(sheet, rows, "name")
                tries = 0
            elif isinstance(value, str) and value.lower() == "inputs":
                rows = self._read_tree_parameter(sheet, rows, "input")
                tries = 0
            elif isinstance(value, str) and value.lower() == "outputs":
                rows = self._read_tree_parameter(sheet, rows, "output")
                tries = 0
            elif isinstance(value, str) and value.lower() == "weights":
                rows = self._read_tree_parameter(sheet, rows, "weight")
                tries = 0
            else:
                rows = rows + 1
                tries = tries + 1                

    def _read_profile(self, sheet=None):
        ''' This function reads the information of different profiles for different parameters

            Parameters
            ----------
            Mandatory:\\
            sheet : excel sheet to be read
        '''
        rows = 2
        tries = 0

        typ=None
        subtype=None
        position_tree={}
        name=None
        ID=None
        group=None
        zone=None

        valid_types = ["bus", "branch", "generator"]
        parameter_elements = self.parameters_names_nodes.copy()
        parameter_elements.update(self.parameters_names_branches)
        parameter_elements.update(self.parameters_names_generators)

        while tries < 20:
            value = sheet.cell(row=rows, column=1).value
            if isinstance(value, str):
                value = value.replace(u'\xa0',' ')
            if isinstance(value, str) and self.parameters_profile.get(value.lower()) == "type":
                value = sheet.cell(row=rows, column=2).value
                rows = rows + 1
                if isinstance(value, str) and value.lower() in valid_types:
                    typ = value.lower()
                else:
                    print("WARNING! Type {} does not exist in the list of valid types".format(value))
            elif typ=="branch" and isinstance(value, str) and self.parameters_profile.get(value.lower()) == "subtype":
                value = sheet.cell(row=rows, column=2).value
                rows = rows + 1
                subtype = value.lower()
                # if isinstance(value, str) and self.accepted_types_branches.get(value.lower(), None) is not None:
                #     subtype = value.lower()
                # else:
                #     print("WARNING! Subtype {} does not exist in the list of valid subtypes for branches".format(value))
            elif typ=="generator" and isinstance(value, str) and self.parameters_profile.get(value.lower()) == "subtype":
                value = sheet.cell(row=rows, column=2).value
                rows = rows + 1
                subtype = value.lower()
                # if isinstance(value, str) and self.accepted_types_generators.get(value.lower(), None) is not None:
                #     subtype = value.lower()
                # else:
                #     print("WARNING! Subtype {} does not exist in the list of valid subtypes for generators".format(value))
            elif typ=="bus" and isinstance(value, str) and self.parameters_profile.get(value.lower()) == "subtype" and sheet.cell(row=rows, column=2).value is not None:
                rows = rows + 1
                print("WARNING! buses does not have any subtype")
            elif isinstance(value, str) and self.parameters_profile.get(value.lower()) == "pt":
                pos = 1
                cols = 2
                value = sheet.cell(row=rows, column=cols).value
                while value is not None:
                    position_tree[str(pos)] = value.lower()
                    pos = pos + 1
                    cols = cols + 1
                    value = sheet.cell(row=rows, column=cols).value
                rows = rows + 1
            elif isinstance(value, str) and self.parameters_profile.get(value.lower()) == "name":
                value = sheet.cell(row=rows, column=2).value
                rows = rows + 1
                if isinstance(value, str) and parameter_elements.get(value.lower()) is not None:
                    name = parameter_elements.get(value.lower())
                else:
                    print("WARNING! name {} is not a valid name".format(value))
            elif isinstance(value, str) and self.parameters_profile.get(value.lower()) == "group":
                value = sheet.cell(row=rows, column=2).value
                rows = rows + 1
                if value is not None:
                    group = value
            elif isinstance(value, str) and self.parameters_profile.get(value.lower()) == "zone":
                value = sheet.cell(row=rows, column=2).value
                rows = rows + 1
                if value is not None:
                    zone = value
            elif (isinstance(value, str) and value.lower() == "dataset") and (isinstance(sheet.cell(row=rows, column=2).value, str) and sheet.cell(row=rows, column=2).value.lower() == "hour") and (isinstance(sheet.cell(row=rows+1, column=1).value, str) and sheet.cell(row=rows+1, column=1).value.lower() == "id"):
                hour_row = rows + 1
                rows = rows + 2
                cols = 1
                value = sheet.cell(row=rows, column=cols).value
                tries = 0
                while value is not None and name is not None:
                    ID=value
                    cols = cols + 1
                    time = 0
                    value = sheet.cell(row=rows, column=cols).value
                    while value is not None:
                        self.profiles_info.append(profile_parameter(name=name, position_tree=position_tree, hour= sheet.cell(row=hour_row, column=cols).value if sheet.cell(row=hour_row, column=cols).value is not None else time, ID=ID, type=typ, subtype=subtype, value=value, group=group, zone=zone))
                        cols = cols + 1
                        time = time + 1
                        value = sheet.cell(row=rows, column=cols).value
                    rows = rows + 1
                    cols = 1
                    value = sheet.cell(row=rows, column=cols).value
                typ=None
                subtype=None
                position_tree={}
                name=None
                ID=None
                group=None
                zone=None
            else:
                rows = rows + 1
                tries = tries + 1

    def _read_energy_connections(self, model=None, sheet=None):
        ''' This function reads the information of connections between variables for the models in pyene

            Parameters
            ----------
            Mandatory:\\
            model : python object containing the mathematical model and all parameters\\
            sheet : excel sheet to be read
        '''

        names_list = copy.copy(self.variables_balance_tree)
        names_list.extend(self.variables_OPF)

        tries = 0
        rows = 2
        while tries < 20:
            info = information()
            value = sheet.cell(row=rows, column=1).value
            if isinstance(value, str):
                value = value.replace(u'\xa0',' ')
            if isinstance(value, str) and value == "begin":
                tries = 0
                while value != "end":
                    rows += 1
                    value = sheet.cell(row=rows, column=1).value
                    if isinstance(value, str):
                        value = value.replace(u'\xa0',' ')
                    if isinstance(value, str) and (self.all_parameters.get(value.lower()) is not None or value.lower() in names_list):
                        original_name = value
                        if self.all_parameters.get(value.lower()) is not None:
                            name = self.all_parameters.get(value.lower())
                        if value.lower() in names_list:
                            name = value.lower()
                        col = 1
                        if self.all_list_parameters.get(value.lower()) is not None:
                            val = []
                            while value is not None:
                                col += 1
                                value = sheet.cell(row=rows, column=col).value
                                if isinstance(value, str):
                                    value = value.replace(u'\xa0',' ')
                                    if self.all_parameters.get(value.lower()) is not None:
                                        value = self.all_parameters.get(value.lower())
                                    if value.lower() in names_list:
                                        value = value.lower()
                                if value is not None:
                                    val.append(value)
                        elif self.all_unique_parameters.get(value.lower()) is not None:
                            value = sheet.cell(row=rows, column=2).value
                            if isinstance(value, str):
                                value = value.replace(u'\xa0',' ')
                                if self.all_parameters.get(value.lower()) is not None:
                                    value = self.all_parameters.get(value.lower())
                                elif value.lower() in names_list:
                                    value = value.lower()
                                elif value.lower() == "all":
                                    value = value.lower()
                            val = value
                        if not val:
                            val = None
                        data_type = self._check_data_type(original_name)
                        if data_type is not None and val is not None:
                            info.characteristics.append(characteristic(name=name, value=val, data_type=data_type))
                if info.characteristics:
                    info.characteristics.append(characteristic(name="name", value="connection", data_type="string"))
                    info.value = True
                    info.data_type = "bool"
                    self.data.connections.append(info)
            else:
                rows = rows + 1
                tries = tries + 1
        
        for connection in self.data.connections:
            if connection.get_characteristic("ID") == "all" and connection.exist("type"):
                list_IDs = []
                for param in model.data.data[connection.get_characteristic("type")]:
                    if param.get_characteristic("ID") not in list_IDs and ( (connection.exist("subtype") and param.exist("subtype") and all(item in param.get_characteristic("subtype") for item in connection.get_characteristic("subtype")) ) or (connection.exist("group") and param.exist("group") and param.get_characteristic("group") == connection.get_characteristic("group")) or (connection.exist("zone") and param.exist("zone") and param.get_characteristic("zone") == connection.get_characteristic("zone")) or (not connection.exist("subtype") and not connection.exist("group") and not connection.exist("zone"))):
                        list_IDs.append(param.get_characteristic("ID"))                            

                for id in list_IDs:
                    new_connection = copy.deepcopy(connection)
                    new_connection.update_characteristic("ID", id)
                    model.data.connections.append(new_connection)

    def _read_functions(self, sheet=None):
        ''' This function reads the information of mathematical functions (polynomial or piecewise) for parameters.

            Parameters
            ----------
            Mandatory:\\
            model : python object containing the mathematical model and all parameters\\
            sheet : excel sheet to be read
        '''
        tries = 0
        rows = 2
        while tries < 20:
            info = information()
            value = sheet.cell(row=rows, column=1).value
            if isinstance(value, str):
                value = value.replace(u'\xa0',' ')
            if isinstance(value, str) and value == "begin":
                tries = 0
                param_names = []
                param_origin_names = []
                initial = True
                while value != "end":
                    rows += 1
                    col = 0
                    while value is not None and initial:
                        col += 1
                        value = sheet.cell(row=rows, column=col).value
                        if isinstance(value, str):
                            value = value.replace(u'\xa0',' ')
                            if self.all_parameters.get(value.lower()) is not None:
                                original_name = value
                                value = self.all_parameters.get(value.lower())
                        if value is not None:
                            param_origin_names.append(original_name)
                            param_names.append(value)
                    if len(param_names) == 0:
                        print ("WARNING! Provided parameter or characteristic name *{}* in excel sheet in column 1 and row {} is not a valid name".format(value, rows))
                        return
                    elif len(param_names) == 1:
                        initial = False
                        value = sheet.cell(row=rows, column=1).value
                        if value == "end":
                            break
                        if isinstance(value, str):
                            value = value.replace(u'\xa0',' ')
                        if self.all_unique_parameters.get(value.lower()) is not None:
                            na = self.all_parameters.get(value.lower())
                            value = sheet.cell(row=rows, column=2).value
                            data_type = self._check_data_type(value.lower())
                            info.characteristics.append(characteristic(name=na, value=value, data_type=data_type))
                        elif self.all_list_parameters.get(value.lower()) is not None:
                            na = self.all_parameters.get(value.lower())
                            col = 1
                            values = []
                            while value is not None:
                                col += 1
                                value = sheet.cell(row=rows, column=col).value
                                if isinstance(value, str):
                                    value = value.replace(u'\xa0',' ')
                                    if self.all_parameters.get(value.lower()) is not None:
                                        value = self.all_parameters.get(value.lower())
                                if value is not None:
                                    values.append(value)
                            info.characteristics.append(characteristic(name="name", value=na, data_type="string"))
                            info.characteristics.append(characteristic(name="coefficients", value=value[nu:], data_type=data_type))
                            info.value = True
                            info.data_type = "bool"
                    else:
                        col = 0
                        values = []
                        if initial:
                            rows += 1
                            initial = False
                        value = sheet.cell(row=rows, column=1).value
                        if value == "end":
                            break
                        while value is not None:
                            col += 1
                            value = sheet.cell(row=rows, column=col).value
                            if isinstance(value, str):
                                value = value.replace(u'\xa0',' ')
                                if self.all_parameters.get(value.lower()) is not None:
                                    value = self.all_parameters.get(value.lower())
                            if value is not None:
                                values.append(value)
                        for nu, (o_na, na) in enumerate(zip(param_origin_names, param_names)):
                            data_type = self._check_data_type(o_na)
                            if self.all_unique_parameters.get(o_na.lower()) is not None:
                                info.characteristics.append(characteristic(name=na, value=values[nu], data_type=data_type))
                            elif self.all_list_parameters.get(o_na.lower()) is not None:
                                info.characteristics.append(characteristic(name="name", value=na, data_type="string"))
                                info.characteristics.append(characteristic(name="coefficients", value=values[nu:], data_type=data_type))
                                info.value = True
                                info.data_type = "bool"
                                if (nu+1 != len(param_names)):
                                    print("WARNING! Inconsisten data for function *{}*. The name and list of coefficient should be after pieces".format(o_na))
                                    return
                                self.data.functions.append(info)
                                info = information()
                if len(param_names) == 1:
                    self.data.functions.append(info)
            else:
                rows = rows + 1
                tries = tries + 1

    def _read_output_requirements(self, model=None, sheet=None):
        ''' This function reads the information of outputs for pyene results

            Parameters
            ----------
            Mandatory:\\
            model : python object containing the mathematical model and all parameters\\
            sheet : excel sheet to be read
        '''
        names_list = copy.copy(self.variables_balance_tree)
        names_list.extend(self.variables_OPF)

        tries = 0
        rows = 2
        while tries < 20:
            info = information()
            value = sheet.cell(row=rows, column=1).value
            if isinstance(value, str):
                value = value.replace(u'\xa0',' ')
            if isinstance(value, str) and value == "begin":
                tries = 0
                while value != "end":
                    rows += 1
                    value = sheet.cell(row=rows, column=1).value
                    if isinstance(value, str):
                        value = value.replace(u'\xa0',' ')
                    if isinstance(value, str) and (self.all_parameters.get(value.lower()) is not None or value.lower() in names_list):
                        original_name = value
                        if self.all_parameters.get(value.lower()) is not None:
                            name = self.all_parameters.get(value.lower())
                        if value.lower() in names_list:
                            name = value.lower()
                        col = 1
                        val = []
                        while value is not None:
                            col += 1
                            value = sheet.cell(row=rows, column=col).value
                            if isinstance(value, str):
                                value = value.replace(u'\xa0',' ')
                                if self.all_parameters.get(value.lower()) is not None:
                                    value = self.all_parameters.get(value.lower())
                                elif value is not None:
                                    value = value.lower()
                            if value is not None:
                                val.append(value)
                        if not val:
                            val = None
                        if val and len(val) == 1 and (val[0] == "all" or val[0] in names_list):
                            data_type = "v_string"
                        elif val:
                            data_type = self._check_data_type(original_name)
                            if (data_type[0:2] != "v_"):
                                data_type = "v_" + data_type
                        else:
                            data_type = None
                        if data_type is not None and val is not None:
                            info.characteristics.append(characteristic(name=name, value=val, data_type=data_type))
                if info.characteristics:
                    info.value = True
                    info.data_type = "bool"
                    self.data.outputs.append(info)
            else:
                rows = rows + 1
                tries = tries + 1
        model.data.outputs = self.data.outputs

    def _check_data_type(self, name=None):
        ''' This function the type of data that is related to the name

            Parameters
            ----------
            Mandatory:\\
            name : name of parameter
            
            Return
            ----------
            This functions returns the type of data related to the name.\\
            return data_type (str)
        '''
        if self.parameters_integer.get(name.lower()) is not None:
            return "integer"
        elif self.parameters_double.get(name.lower()) is not None:
            return "double"
        elif self.parameters_string.get(name.lower()) is not None:
            return "string"
        elif self.parameters_bool.get(name.lower()) is not None:
            return "bool"
        elif self.parameters_list_integer.get(name.lower()) is not None:
            return "v_integer"
        elif self.parameters_list_double.get(name.lower()) is not None:
            return "v_double"
        elif self.parameters_list_string.get(name.lower()) is not None:
            return "v_string"
        elif self.parameters_list_bool.get(name.lower()) is not None:
            return "v_bool"
        else:
            return None
                    
    def _remove_uncorrelated(self, parameters, name_comp, value_comp, list_pos):
        list_remove = []
        for num, pos in enumerate(list_pos):
            if (name_comp == "subtype" and parameters[pos].exist("subtype") and value_comp not in parameters[pos].get_characteristic("subtype")) or (name_comp != "subtype" and parameters[pos].exist(name_comp) and value_comp not in parameters[pos].get_characteristic(name_comp)) or (name_comp == "subtype" and parameters[pos].exist("subtype") and value_comp in parameters[pos].get_characteristic("subtype") and (parameters[pos].exist("pt") or parameters[pos].exist("hour"))) or (name_comp != "subtype" and parameters[pos].exist(name_comp) and value_comp in parameters[pos].get_characteristic(name_comp) and (parameters[pos].exist("pt") or parameters[pos].exist("hour"))):
                list_remove.append(num)
        list_pos = [ele for num,ele in enumerate(list_pos) if num not in list_remove]
        return list_pos

    def _extract_parameters_correlation(self, network_info=None, profile=None, delete_only=False):
        ''' This function extract all parameters in the list of network parameters in model that have the same identification characteristics that the profile

            Parameters
            ----------
            Mandatory:\\
            model : python object containing the mathematical model and all parameters\\
            profile : profile object used for comparison of information
        '''
        if profile.ID == "all":
            list_pos = []
            # getting parameters indexes to be correlated
            for num, parameter in enumerate(network_info.data[profile.type]):
                if parameter.get_characteristic("name") == profile.name:
                    list_pos.append(num)
            # Removing the IDs with different subtype that the specified subtype
            if profile.subtype is not None:
                list_pos = self._remove_uncorrelated(network_info.data[profile.type], "subtype", profile.subtype, list_pos)
            if profile.group is not None:
                list_pos = self._remove_uncorrelated(network_info.data[profile.type], "group", profile.group, list_pos)
            if profile.zone is not None:
                list_pos = self._remove_uncorrelated(network_info.data[profile.type], "zone", profile.zone, list_pos)
            # extracting parameters
            if not delete_only:
                params = []
                for pos in list_pos:
                    params.append(network_info.data[profile.type][pos])
                return np.array(params)
            else:
                list_pos = [pos for pos in list_pos if not network_info.data[profile.type][pos].exist("pt") and not network_info.data[profile.type][pos].exist("hour")]
                params = np.array([ele for num,ele in enumerate(network_info.data[profile.type]) if num not in list_pos])
                network_info.data[profile.type] = params
                return network_info
        else:
            if not delete_only:
                for num, parameter in enumerate(network_info.data[profile.type]):
                    if parameter.get_characteristic("ID") == profile.ID and parameter.get_characteristic("name") == profile.name:
                        return num
            else:
                num_erase = -1
                for num, parameter in enumerate(network_info.data[profile.type]):
                    if parameter.get_characteristic("ID") == profile.ID and parameter.get_characteristic("name") == profile.name:
                        num_erase = num
                        break
                if num_erase == -1:
                    raise ValueError("value not found")
                network_info.data[profile.type] = np.delete(network_info.data[profile.type], num_erase)
                return network_info

    def _extract_profiles_correlation(self, prof_compare=None):
        ''' This function extract all profiles that have the same identification characteristics that prof_compare

            Parameters
            ----------
            Mandatory:\\
            prof_compare : profile object used for comparison of information
        '''    
        list_pos = []
        profs = []
        for num, profile in enumerate(self.profiles_info):
            if prof_compare.type == profile.type and prof_compare.subtype == profile.subtype and prof_compare.name == profile.name and prof_compare.ID == profile.ID:
                list_pos.append(num)
        for pos in list_pos:
            profs.append(self.profiles_info[pos])
        self.profiles_info = [ele for num,ele in enumerate(self.profiles_info) if num not in list_pos]
        return profs

    def _check_profile_existance(self, model=None, profile=None):
        ''' This checks if the profiles were already created and the function erase them

            Parameters
            ----------
            Mandatory:\\
            model : python object containing the mathematical model and all parameters\\
            profile : profile object used for comparison of information
        '''
        exist = False
        list_remove = []
        for pos, parameter in enumerate(model.data.data[profile.type]):
            if parameter.get_characteristic("name") == profile.name and ((parameter.exist("pt") and parameter.get_characteristic("pt") == profile.position_tree and parameter.exist("hour") and parameter.get_characteristic("hour") == profile.hour) or (not parameter.exist("pt") and parameter.exist("hour") and parameter.get_characteristic("hour") == profile.hour)):
                exist = True
                list_remove.append(pos)
                if profile.subtype is None:
                    print("WARNING! parameter {} has been already set with the representative day {} and hour {}.".format(parameter.name, parameter.position_tree, parameter.hour))

        if profile.subtype is not None:
            exist = False
            new_list_remove = []
            for remove in list_remove:
                ID = model.data.data[profile.type][remove].get_characteristic("ID")
                if profile.subtype in model.data.data[profile.type][remove].get_characteristic("subtype"):
                    exist = True
                    new_list_remove.append(pos)
                    print("WARNING! parameter {} has been already set with the representative day {} and hour {}.".format(model.data.data[profile.type][remove].get_characteristic("name"), model.data.data[profile.type][remove].get_characteristic("pt"), model.data.data[profile.type][remove].get_characteristic("hour")))
            list_remove = new_list_remove
        if exist:
            print("The values will be updated with the most up-to-date information. Check your profile inputs to avoid duplications or errors.")
        # Removing parameters from model
        model.data.data[profile.type] = np.array([ele for num,ele in enumerate(model.data.data[profile.type]) if num not in list_remove])

        return exist, model
    
    def _correlate_information(self, model=None, network_info=None):
        ''' This function correlates the all profiles and the respective parameters

            Parameters
            ----------
            Mandatory:\\
            model : python object containing the mathematical model and all parameters
        '''
        # Correlate profiles with parameters
        correlate_profiles = False
        for opt in model.model_options:
            if opt.name == "representative periods" and opt.value:
                correlate_profiles = True
        counter_prof = 0
        while self.profiles_info and correlate_profiles:
            params = self._extract_parameters_correlation(network_info, self.profiles_info[counter_prof])
            exist, model = self._check_profile_existance(model, self.profiles_info[counter_prof])
            if not exist:
                model.data = self._extract_parameters_correlation(model.data, self.profiles_info[counter_prof], True)
            profs = self._extract_profiles_correlation(self.profiles_info[counter_prof])
            if isinstance(params, np.ndarray):
                new_list = []
                for param in params:
                    for prof in profs:
                        new_info = information()
                        new_info.characteristics = copy.copy(param.characteristics)
                        tree = []
                        for _, value in prof.position_tree.items():
                            tree.append(value)
                        new_info.characteristics.append(characteristic(name="pt", value=tree, data_type="v_string"))
                        new_info.characteristics.append(characteristic(name="hour", value=prof.hour, data_type="double"))
                        new_info.data_type = param.data_type
                        new_info.value = param.value * prof.value
                        new_list.append(new_info)
                model.data.data[profs[0].type] = np.append(model.data.data[profs[0].type], new_list)
            else:
                new_list = []
                for prof in profs:
                    new_info = information()
                    new_info.characteristics = copy.copy(network_info.data[prof.type][params].characteristics)
                    tree = []
                    for _, value in prof.position_tree.items():
                        tree.append(value)
                    new_info.characteristics.append(characteristic(name="pt", value=tree,data_type="v_string"))
                    new_info.characteristics.append(characteristic(name="hour", value=prof.hour,data_type="double"))
                    new_info.data_type = network_info.data[prof.type][params].data_type
                    new_info.value = network_info.data[prof.type][params].value * prof.value
                    new_list.append(new_info)
                model.data.data[profs[0].type] = np.append(model.data.data[profs[0].type], new_list)

        # Correlate characteristics with functions information
        for info in self.data.functions:
            for _, parameters in model.data.data.items():
                found = False
                for parameter in parameters:
                    if parameter.get_characteristic("ID") == info.get_characteristic("ID"):
                        chars = [cha for cha in parameter.characteristics if cha.name != "ID" and cha.name != "name"]
                        info.characteristics.extend(chars)
                        found = True
                        break
                if found:
                    break
        model.data.functions = self.data.functions

        #     for parameter in model.network_parameters:
        #         if parameter.ID == info.get_characteristic("ID"):
        #             info.characteristics.append(characteristic(name="type", value=parameter.type, data_type="string"))
        #             subtype = []
        #             for par in model.network_parameters:
        #                 if par.name == "subtype" and parameter.ID == par.ID and parameter.type == par.type and par.value is not None:
        #                     subtype.append(par.value)
        #                 elif par.name != "subtype" and par.name in set(self.parameters_integer.values()) and parameter.ID == par.ID and parameter.type == par.type and par.value is not None:
        #                     info.characteristics.append(characteristic(name=par.name, value=par.value, data_type="integer"))
        #                 elif par.name != "subtype" and par.name in set(self.parameters_string.values()) and parameter.ID == par.ID and parameter.type == par.type and par.value is not None:
        #                     info.characteristics.append(characteristic(name=par.name, value=str(par.value), data_type="string"))
        #             if subtype:
        #                 info.characteristics.append(characteristic(name="subtype", value=subtype, data_type="v_string"))
        #             break
        # model.data.functions = self.data.functions

    def _load_balance_tree(self, model=None):
        ''' This function load the information of the balance tree in the model

            Parameters
            ----------
            Mandatory:\\
            model : python object containing the mathematical model and all parameters
        '''
        names = [param for param in self.tree_info if param.name == "name"]
        self.tree_info = [param for param in self.tree_info if param.name != "name"]
        for info in self.tree_info:
            name_pos = ""
            for name in names:
                if name.level == info.level and name.pos == info.pos:
                    name_pos = name.value
                    break
            model.tree_parameters.append(tree_parameters(name=info.name, level=info.level, name_node=name_pos, value=info.value))
        self.tree_info = []


    def read_excel(self, energy_file=None, model=None, **kwargs):
        """Load variables, parameters and options to solve the specified energy 
        optimisation problems with the multi-objective optimisation algorithm

        Parameters
        ----------
        Mandatory:\\
        energy_file : Path to file containing options for energy optimisation

        """
        self.tree_info = []
        self.profiles_info = []
        self.data = parameters()
        self.files2open = [energy_file]
        while self.files2open:
            workbook = load_workbook(filename=self.files2open.pop())
            for ws in workbook.worksheets:
                value = ws.cell(row=1, column=1).value.replace(u'\xa0',' ')
                if value == "model":
                    self._read_model_options(model, ws)
                elif value == "bus":
                    self._read_info_network(model, ws, self.parameters_names_nodes, "bus")
                elif value == "branch":
                    self._read_info_network(model, ws, self.parameters_names_branches, "branch")
                elif value == "generator":
                    self._read_info_network(model, ws, self.parameters_names_generators, "generator")
                elif value == "tree":
                    self._read_info_tree(ws)
                elif value == "profile":
                    self._read_profile(ws)
                elif value == "connections energy":
                    self._read_energy_connections(model, ws)
                elif value == "functions":
                    self._read_functions(ws)
                elif value == "outputs":
                    self._read_output_requirements(model, ws)
                else:
                    print("WARNING! option {} not identified".format(value))
        
        self._load_balance_tree(model)

        network_info = copy.deepcopy(model.data)
        self._correlate_information(model, network_info)
